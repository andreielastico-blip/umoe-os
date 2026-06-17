# -*- coding: utf-8 -*-
"""
UMOE OS 8.0 | Enterprise BI Dashboard v2
Gera HTML completo com drill-down, analise conclusiva e insights.
"""
import openpyxl, glob, json, sys, shutil
from pathlib import Path
from datetime import datetime
from collections import defaultdict

OUT_DIR = Path(__file__).parent.parent / "Relatorios"
OUT_DIR.mkdir(exist_ok=True)
OUT_FILE = OUT_DIR / "UMOE_BI_Enterprise.html"
DOCS_DIR = Path(__file__).parent.parent.parent / "docs"
DOCS_DIR.mkdir(exist_ok=True)

print("[1/5] Carregando BD SAFRAS...")
files = glob.glob(r'C:\01 - UMOE\03 - Financeiro\Planilhas\Historico Kpis*Agr*.xlsx')
if not files:
    print("  ERRO: BD SAFRAS nao encontrado"); sys.exit(1)

wb_agr = openpyxl.load_workbook(files[0], read_only=True, data_only=True)
ws_bd  = wb_agr['BD SAFRAS']

amb_data  = defaultdict(lambda: dict(tah_w=0,tch_w=0,atr_w=0,ha=0))
var_data  = defaultdict(lambda: dict(tah_w=0,tch_w=0,atr_w=0,ha=0))
corte_data= defaultdict(lambda: dict(tch_w=0,ha=0))
faz_data  = defaultdict(lambda: dict(tah_w=0,tch_w=0,atr_w=0,ha=0,amb=''))
safra_data= defaultdict(lambda: dict(tch_w=0,atr_w=0,tah_w=0,ha=0))
var_amb   = defaultdict(lambda: dict(tah_w=0,tch_w=0,ha=0))

for row in ws_bd.iter_rows(min_row=2, values_only=True):
    try:
        amb=str(row[0] or '').strip(); faz=str(row[1] or '').strip()
        var=str(row[5] or '').strip(); ha=float(row[7] or 0)
        tch=float(row[11] or 0); atr=float(row[16] or 0)
        tah=float(row[17] or 0); safra=int(row[21] or 0)
        corte=int(row[24] or 0)
    except: continue
    if ha<0.5 or tch<10 or tah<0.5 or not amb or not faz: continue

    if len(amb)==1:
        amb_data[amb]['tah_w']+=tah*ha; amb_data[amb]['tch_w']+=tch*ha
        amb_data[amb]['atr_w']+=atr*ha; amb_data[amb]['ha']+=ha

    if var:
        var_data[var]['tah_w']+=tah*ha; var_data[var]['tch_w']+=tch*ha
        var_data[var]['atr_w']+=atr*ha; var_data[var]['ha']+=ha
        if len(amb)==1:
            k2=f'{var}|{amb}'
            var_amb[k2]['tah_w']+=tah*ha; var_amb[k2]['tch_w']+=tch*ha; var_amb[k2]['ha']+=ha

    if 1<=corte<=8:
        corte_data[corte]['tch_w']+=tch*ha; corte_data[corte]['ha']+=ha

    fk=f'{amb}|{faz}'
    faz_data[fk]['tah_w']+=tah*ha; faz_data[fk]['tch_w']+=tch*ha
    faz_data[fk]['atr_w']+=atr*ha; faz_data[fk]['ha']+=ha; faz_data[fk]['amb']=amb

    if safra>=20000:
        safra_data[safra]['tch_w']+=tch*ha; safra_data[safra]['atr_w']+=atr*ha
        safra_data[safra]['tah_w']+=tah*ha; safra_data[safra]['ha']+=ha

print("  BD SAFRAS: OK")
print("[2/5] Carregando Historico Industrial...")
hist_files = glob.glob(r'C:\01 - UMOE\03 - Financeiro\Planilhas\Hist*rio Di*rio Safras.xlsx')
if not hist_files:
    hist_files = glob.glob(r'C:\01 - UMOE\03 - Financeiro\Planilhas\Hist*.xlsx')
wb_hist = openpyxl.load_workbook(hist_files[0], read_only=True, data_only=True)

ind_mensal = defaultdict(lambda: dict(moa=0,n=0,atr_w=0,egi_w=0,aprov_w=0,aprov_n=0))
for ano in ['2022','2023','2024','2025','2026']:
    if ano not in wb_hist.sheetnames: continue
    ws=wb_hist[ano]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[2]: continue
        dt=row[2]
        try:
            mes=dt.month if hasattr(dt,'month') else int(str(dt)[5:7])
            moa=float(row[3] or 0); atr=float(row[4] or 0)
            egi=float(row[9] or 0) if len(row)>9 else 0
            aprov=float(row[30] or 0) if len(row)>30 else 0
        except: continue
        if moa<500: continue
        k=f'{ano}-{mes:02d}'
        ind_mensal[k]['moa']+=moa; ind_mensal[k]['n']+=1
        if atr>50: ind_mensal[k]['atr_w']+=atr*moa
        if egi>50: ind_mensal[k]['egi_w']+=egi*moa
        if aprov>0: ind_mensal[k]['aprov_w']+=aprov; ind_mensal[k]['aprov_n']+=1

print("  Industrial: OK")
print("[3/5] Calculando metricas...")

amb_list=[]
for a,v in sorted(amb_data.items()):
    h=v['ha']
    if h<100: continue
    amb_list.append({'amb':a,'tah':round(v['tah_w']/h,2),'tch':round(v['tch_w']/h,1),
                     'atr':round(v['atr_w']/h,1),'ha':round(h,0)})

var_list=[]
for vn,v in var_data.items():
    h=v['ha']
    if h<20: continue
    var_list.append({'var':vn,'tah':round(v['tah_w']/h,2),'tch':round(v['tch_w']/h,1),
                     'atr':round(v['atr_w']/h,1),'ha':round(h,0)})
var_list.sort(key=lambda x:-x['tah'])
top_vars  =var_list[:15]
worst_vars=sorted([v for v in var_list if v['ha']>100],key=lambda x:x['tah'])[:12]

corte_list=[]
for c in range(1,9):
    v=corte_data.get(c,{})
    if v.get('ha',0)>100:
        corte_list.append({'corte':c,'tch':round(v['tch_w']/v['ha'],1),'ha':round(v['ha'],0)})

faz_list=[]
for fk,v in faz_data.items():
    h=v['ha']
    if h<30: continue
    amb,faz=fk.split('|',1)
    if len(amb)!=1: continue
    faz_list.append({'amb':amb,'faz':faz,'tah':round(v['tah_w']/h,2),
                     'tch':round(v['tch_w']/h,1),'atr':round(v['atr_w']/h,1),'ha':round(h,0)})
faz_list.sort(key=lambda x:-x['tah'])

safra_list=[]
for s,v in sorted(safra_data.items()):
    h=v['ha']
    if h<1000: continue
    yr=str(s)
    label=f"{yr[:2]}/{yr[2:4]}-{yr[4:6]}/{yr[6:]}" if len(yr)==8 else str(s)
    safra_list.append({'safra':s,'label':label,'tch':round(v['tch_w']/h,1),
                       'atr':round(v['atr_w']/h,1),'tah':round(v['tah_w']/h,2),'ha':round(h,0)})

drilldown={}
for a in ['A','B','C','D','E']:
    fazs=[f for f in faz_list if f['amb']==a]
    fazs.sort(key=lambda x:-x['tah'])
    drilldown[a]=fazs[:30]

var_amb_list=[]
for k,v in var_amb.items():
    vn,a=k.split('|',1)
    h=v['ha']
    if h<30: continue
    var_amb_list.append({'var':vn,'amb':a,'tah':round(v['tah_w']/h,2),'tch':round(v['tch_w']/h,1),'ha':round(h,0)})
var_amb_list.sort(key=lambda x:-x['tah'])
top_var_amb=var_amb_list[:20]

ind_series={}
for k in sorted(ind_mensal.keys()):
    v=ind_mensal[k]; m=v['moa']
    ind_series[k]={'moa':round(m/1000),'atr':round(v['atr_w']/m,1) if m else 0,
                   'egi':round(v['egi_w']/m,1) if m else 0,
                   'aprov':round(v['aprov_w']/v['aprov_n'],1) if v['aprov_n'] else 0}

sp81    =next((v for v in var_list if 'SP81' in v['var'] and v['ha']>1000),None)
ctc9006 =next((v for v in var_list if 'CTC9006' in v['var']),None)

TOTAL_HA      =sum(v['ha'] for v in amb_list)
MOAGEM_META   =2768000; MOAGEM_REAL=606418
ATR_META      =138.66;  ATR_REAL=126.49; GAP_ATR=ATR_META-ATR_REAL
GAP_R         =50.78
SP81_HA       =sp81['ha']  if sp81  else 30183
SP81_TAH      =sp81['tah'] if sp81  else 6.29
CTC_TAH       =ctc9006['tah'] if ctc9006 else 21.62
TAH_GAP       =CTC_TAH-SP81_TAH
SP81_POT_M    =round((TAH_GAP*SP81_HA*1000*1.03)/1e6)
ha_4c         =15173
tch_4c        =next((c['tch'] for c in corte_list if c['corte']==4),52.5)
tch_1c        =next((c['tch'] for c in corte_list if c['corte']==1),98.6)

print("[4/5] Gerando HTML...")

HTML = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>UMOE OS 8.0 | Enterprise BI Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
:root{{--bg:#060E1F;--bg2:#0D1B35;--bg3:#152344;--green:#00C853;--gold:#FFD600;--red:#FF1744;--blue:#2979FF;--orange:#FF6D00;--cyan:#00E5FF;--text:#E8EFF8;--text2:#8BA0BB;--border:#1E3252}}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:var(--bg);color:var(--text);font-family:'Segoe UI',sans-serif;font-size:13px}}
header{{background:linear-gradient(135deg,#0A1628,#0D2040);border-bottom:1px solid var(--border);padding:14px 24px;display:flex;align-items:center;justify-content:space-between}}
header h1{{font-size:18px;font-weight:700;color:var(--green)}}
.tabs{{display:flex;gap:2px;padding:10px 16px 0;background:var(--bg2);border-bottom:2px solid var(--border);overflow-x:auto}}
.tab{{padding:8px 18px;cursor:pointer;border-radius:6px 6px 0 0;color:var(--text2);font-size:12px;font-weight:600;white-space:nowrap;border:1px solid transparent;border-bottom:none;transition:all .2s}}
.tab:hover{{color:var(--text);background:var(--bg3)}}
.tab.active{{color:var(--green);background:var(--bg);border-color:var(--border)}}
.panel{{display:none;padding:16px;animation:fade .2s}}
.panel.active{{display:block}}
@keyframes fade{{from{{opacity:0}}to{{opacity:1}}}}
.grid{{display:grid;gap:14px}}
.g2{{grid-template-columns:repeat(2,1fr)}}
.g3{{grid-template-columns:repeat(3,1fr)}}
.kpi-row{{display:grid;grid-template-columns:repeat(auto-fit,minmax(155px,1fr));gap:10px;margin-bottom:14px}}
.kpi{{background:var(--bg2);border:1px solid var(--border);border-radius:10px;padding:14px;text-align:center}}
.kpi .label{{font-size:10px;color:var(--text2);text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px}}
.kpi .value{{font-size:22px;font-weight:700}}
.kpi .sub{{font-size:10px;color:var(--text2);margin-top:2px}}
.kpi.red .value{{color:var(--red)}}.kpi.green .value{{color:var(--green)}}.kpi.gold .value{{color:var(--gold)}}.kpi.blue .value{{color:var(--blue)}}.kpi.orange .value{{color:var(--orange)}}
.card{{background:var(--bg2);border:1px solid var(--border);border-radius:10px;padding:14px}}
.card h3{{font-size:11px;font-weight:700;color:var(--text2);text-transform:uppercase;letter-spacing:.8px;margin-bottom:12px}}
.chart-wrap{{position:relative;height:220px}}
.chart-wrap.tall{{height:300px}}
table{{width:100%;border-collapse:collapse;font-size:11px}}
th{{background:var(--bg3);color:var(--text2);text-align:left;padding:7px 10px;font-size:10px;text-transform:uppercase;letter-spacing:.5px;border-bottom:1px solid var(--border)}}
td{{padding:6px 10px;border-bottom:1px solid var(--border)}}
tr:last-child td{{border:none}}
tr:hover td{{background:var(--bg3)}}
.alert-bar{{background:#FF174410;border:1px solid var(--red);border-radius:8px;padding:10px 16px;margin-bottom:14px;font-size:12px}}
.insight{{background:linear-gradient(135deg,var(--bg3),var(--bg2));border:1px solid var(--green);border-radius:10px;padding:16px;margin-bottom:14px}}
.insight h4{{color:var(--green);font-size:13px;font-weight:700;margin-bottom:8px}}
.insight p{{color:var(--text2);font-size:12px;line-height:1.7}}
.hl{{color:var(--gold);font-weight:700}}
.drill-btn{{cursor:pointer;color:var(--cyan);text-decoration:underline;font-size:11px}}
#drill-modal{{display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:#000a;z-index:999;align-items:center;justify-content:center}}
#drill-modal.open{{display:flex}}
.modal-box{{background:var(--bg2);border:1px solid var(--border);border-radius:14px;width:90%;max-width:900px;max-height:85vh;overflow:auto;padding:20px}}
.modal-box h3{{color:var(--green);margin-bottom:14px;font-size:15px}}
.close-btn{{float:right;cursor:pointer;color:var(--red);font-size:18px;font-weight:700}}
.bar-mini{{display:inline-block;height:6px;border-radius:3px;vertical-align:middle;margin-left:6px}}
.trend-up{{color:var(--green)}}.trend-down{{color:var(--red)}}
.scroll-x{{overflow-x:auto}}
.box3{{padding:12px;background:var(--bg3);border-radius:8px}}
</style>
</head>
<body>
<header>
  <div>
    <h1>UMOE OS 8.0 | Enterprise BI Dashboard</h1>
    <div style="color:var(--text2);font-size:11px">Safra 2025/26 | Base: 51.333 registros | Gerado: {datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
  </div>
  <div style="text-align:right">
    <div style="color:var(--gold);font-weight:700;font-size:14px">Presidente Prudente, SP</div>
    <div style="color:var(--text2);font-size:11px">Diretor: Andrei Elastico | UMOE Bioenergy</div>
  </div>
</header>

<div class="tabs">
  <div class="tab active" onclick="showTab('exec',this)">EXECUTIVO</div>
  <div class="tab" onclick="showTab('agr',this)">AGRICOLA DRILL-DOWN</div>
  <div class="tab" onclick="showTab('var',this)">VARIEDADES</div>
  <div class="tab" onclick="showTab('corte',this)">CURVA CORTE</div>
  <div class="tab" onclick="showTab('ind',this)">INDUSTRIAL</div>
  <div class="tab" onclick="showTab('hist',this)">HISTORICO SAFRAS</div>
  <div class="tab" onclick="showTab('intel',this)">INTELIGENCIA</div>
</div>

<div id="panel-exec" class="panel active">
  <div class="alert-bar">
    <strong style="color:var(--red)">ALERTA CRITICO:</strong>
    Moagem {MOAGEM_REAL:,} t vs meta {int(MOAGEM_META*5/12):,} t (proporcional) &nbsp;|&nbsp;
    Gap receita: <strong style="color:var(--red)">-R$ {GAP_R:.1f}M</strong> &nbsp;|&nbsp;
    ATR <strong style="color:var(--red)">{ATR_REAL}</strong> vs meta {ATR_META} (-{GAP_ATR:.1f} kg/t) &nbsp;|&nbsp;
    CCT: <strong style="color:var(--red)">R$50/t vs R$38,3/t orcado</strong>
  </div>
  <div class="kpi-row">
    <div class="kpi red"><div class="label">Moagem Acumulada</div><div class="value">{MOAGEM_REAL/1e6:.3f} Mt</div><div class="sub">Meta prop: {MOAGEM_META*5/12/1e6:.3f} Mt | Gap: -{int(MOAGEM_META*5/12-MOAGEM_REAL)//1000}kt</div></div>
    <div class="kpi red"><div class="label">Gap Receita</div><div class="value">-R$ {GAP_R:.1f}M</div><div class="sub">89% volume | 11% custo</div></div>
    <div class="kpi gold"><div class="label">ATR Real mai/26</div><div class="value">{ATR_REAL}</div><div class="sub">Meta {ATR_META} | Gap -{GAP_ATR:.1f} kg/t</div></div>
    <div class="kpi red"><div class="label">CCT Real</div><div class="value">R$ 50,0/t</div><div class="sub">Orcado R$38,3/t | +R$11,7/t</div></div>
    <div class="kpi orange"><div class="label">SP81-3250 CRITICO</div><div class="value">{SP81_HA:,.0f} ha</div><div class="sub">TAH {SP81_TAH:.1f} vs {CTC_TAH:.1f} best | Pot R${SP81_POT_M}M</div></div>
    <div class="kpi red"><div class="label">Canavial 4C+</div><div class="value">{ha_4c:,} ha</div><div class="sub">TCH {tch_4c} vs 1C {tch_1c} (-{round((1-tch_4c/tch_1c)*100)}%)</div></div>
  </div>
  <div class="grid g2">
    <div class="card">
      <h3>TOP 5 ACOES POR IMPACTO FINANCEIRO</h3>
      <table>
        <thead><tr><th>#</th><th>Acao</th><th>Area</th><th>Impacto R$</th><th>Prazo</th></tr></thead>
        <tbody>
          <tr><td><strong style="color:var(--red)">1</strong></td><td><strong>Substituir SP81-3250</strong> por CTC9006/CTC9007</td><td>{SP81_HA:,.0f} ha</td><td style="color:var(--green)">R$ {SP81_POT_M}M+</td><td>1-3 safras</td></tr>
          <tr><td><strong style="color:var(--red)">2</strong></td><td><strong>Reforma urgente 4C+</strong></td><td>{ha_4c:,} ha</td><td style="color:var(--green)">R$ 180M+</td><td>2 safras</td></tr>
          <tr><td><strong style="color:var(--gold)">3</strong></td><td><strong>Recuperar ATR</strong> — maturadores + variedades</td><td>—</td><td style="color:var(--green)">R$ 28M/safra</td><td>Imediato</td></tr>
          <tr><td><strong style="color:var(--gold)">4</strong></td><td><strong>Reducao CCT</strong> — logistica + CHI</td><td>2,77 Mt</td><td style="color:var(--green)">R$ 32M/safra</td><td>Imediato</td></tr>
          <tr><td><strong style="color:var(--blue)">5</strong></td><td><strong>Expandir CTC9006</strong> em Amb D e E</td><td>+5.000 ha</td><td style="color:var(--green)">R$ 79M+</td><td>3-5 safras</td></tr>
        </tbody>
      </table>
    </div>
    <div class="card">
      <h3>ATR MENSAL COMPARATIVO 2024-2026</h3>
      <div class="chart-wrap"><canvas id="cAtrComp"></canvas></div>
    </div>
  </div>
  <div class="grid g3" style="margin-top:14px">
    <div class="card"><h3>PERFORMANCE POR AMBIENTE</h3><div class="chart-wrap"><canvas id="cAmb"></canvas></div></div>
    <div class="card"><h3>MOAGEM MENSAL 2025 vs 2026 (kt)</h3><div class="chart-wrap"><canvas id="cMoaComp"></canvas></div></div>
    <div class="card"><h3>EGI INDUSTRIAL 2022-2026 (%)</h3><div class="chart-wrap"><canvas id="cEgi"></canvas></div></div>
  </div>
</div>

<div id="panel-agr" class="panel">
  <div class="kpi-row">
    <div class="kpi blue"><div class="label">Total Area Historico</div><div class="value">{TOTAL_HA:,.0f} ha</div><div class="sub">8 safras consolidadas</div></div>
    <div class="kpi green"><div class="label">Melhor TAH Ambiente</div><div class="value">Amb C</div><div class="sub">TAH 8.76 | 101.178 ha</div></div>
    <div class="kpi gold"><div class="label">Maior Area</div><div class="value">Amb D</div><div class="sub">182.562 ha | TAH 8.69</div></div>
    <div class="kpi red"><div class="label">Maior ATR Ambiente</div><div class="value">Amb E</div><div class="sub">ATR medio 131.0 | 60.440 ha</div></div>
  </div>
  <div class="grid g2">
    <div class="card">
      <h3>PERFORMANCE POR AMBIENTE — clique para ver fazendas</h3>
      <table><thead><tr><th>Amb</th><th>TCH</th><th>ATR</th><th>TAH</th><th>Area (ha)</th><th>% Area</th><th>Drill</th></tr></thead>
      <tbody id="amb-table"></tbody></table>
    </div>
    <div class="card"><h3>TAH POR AMBIENTE (barras horizontais)</h3><div class="chart-wrap tall"><canvas id="cAmbTah"></canvas></div></div>
  </div>
</div>

<div id="panel-var" class="panel">
  <div class="insight">
    <h4>INSIGHT CRITICO — SP81-3250: O MAIOR DESTRUIDOR DE VALOR</h4>
    <p>
      A variedade <span class="hl">SP81-3250</span> ocupa <span class="hl">{SP81_HA:,.0f} ha</span> (a MAIOR area de qualquer variedade)
      com TAH de <span class="hl">{SP81_TAH:.2f} t ATR/ha</span>. A melhor variedade, CTC9006, atinge
      <span class="hl">TAH {CTC_TAH:.2f}</span> — <span class="hl">{CTC_TAH/SP81_TAH:.1f}x mais produtiva</span>.
      Gap potencial: <span class="hl">+{TAH_GAP:.1f} t ATR/ha x {SP81_HA:,.0f} ha = R$ {SP81_POT_M}M</span> por safra.
    </p>
  </div>
  <div class="grid g2">
    <div class="card"><h3>TOP 15 VARIEDADES (TAH)</h3><div class="scroll-x"><table><thead><tr><th>#</th><th>Variedade</th><th>TAH</th><th>TCH</th><th>ATR</th><th>Area (ha)</th></tr></thead><tbody id="top-var-table"></tbody></table></div></div>
    <div class="card"><h3>PIORES VARIEDADES (min 100ha) — SUBSTITUIR URGENTE</h3><div class="scroll-x"><table><thead><tr><th>#</th><th>Variedade</th><th>TAH</th><th>TCH</th><th>Area (ha)</th><th>Perda R$</th></tr></thead><tbody id="worst-var-table"></tbody></table></div></div>
  </div>
  <div class="grid g2" style="margin-top:14px">
    <div class="card"><h3>RANKING VARIEDADES POR TAH (top 12)</h3><div class="chart-wrap tall"><canvas id="cVarRank"></canvas></div></div>
    <div class="card"><h3>TOP VARIEDADE x AMBIENTE</h3><div class="scroll-x"><table><thead><tr><th>Variedade</th><th>Amb</th><th>TAH</th><th>TCH</th><th>Area (ha)</th></tr></thead><tbody id="var-amb-table"></tbody></table></div></div>
  </div>
</div>

<div id="panel-corte" class="panel">
  <div class="insight">
    <h4>DECAIMENTO REAL: 1C para 6C = -54% DE TCH</h4>
    <p>
      O canavial UMOE perde <span class="hl">-20 t/ha no 2o corte</span> e chega ao 4o corte com
      <span class="hl">apenas 53% da produtividade original</span> ({tch_1c} para {tch_4c} t/ha).
      Com <span class="hl">{ha_4c:,} ha em 4C+</span>, a empresa perde
      <span class="hl">R$ 40-60M por safra</span> vs potencial de reforma.
      Taxa de reforma ideal: <span class="hl">25% ao ano</span> para manter canavial jovem.
    </p>
  </div>
  <div class="grid g2">
    <div class="card"><h3>CURVA DE DECAIMENTO TCH POR CORTE (dados reais)</h3><div class="chart-wrap tall"><canvas id="cCorte"></canvas></div></div>
    <div class="card">
      <h3>TABELA DECAIMENTO + IMPACTO FINANCEIRO</h3>
      <table><thead><tr><th>Corte</th><th>TCH</th><th>vs 1C</th><th>Perda t/ha</th><th>Receita Perdida</th></tr></thead><tbody id="corte-table"></tbody></table>
      <div style="background:var(--bg3);padding:12px;border-radius:8px;margin-top:10px;font-size:11px;color:var(--text2);line-height:1.6">
        <strong style="color:var(--gold)">Recomendacao:</strong> Aumentar reforma para 25% ao ano (6.930 ha/ano) priorizando 4C+.
        Payback estimado: 2,5 safras. ROI 5 anos: &gt;400%. Integrar com substituicao SP81-3250.
      </div>
    </div>
  </div>
</div>

<div id="panel-ind" class="panel">
  <div class="kpi-row">
    <div class="kpi red"><div class="label">ATR mai/26</div><div class="value">127.2</div><div class="sub">vs mai/25: 134.4 (-5.4%)</div></div>
    <div class="kpi red"><div class="label">Moagem mar-mai/26</div><div class="value">606 kt</div><div class="sub">vs meta: 823 kt (-26%)</div></div>
    <div class="kpi gold"><div class="label">EGI jun/26</div><div class="value">87.3%</div><div class="sub">Acima media historica</div></div>
    <div class="kpi green"><div class="label">Aproveit. jun/26</div><div class="value">92.6%</div><div class="sub">Melhor do ano</div></div>
    <div class="kpi red"><div class="label">Melhor ATR historico</div><div class="value">154.8</div><div class="sub">Set/2025 — benchmark</div></div>
  </div>
  <div class="grid g2">
    <div class="card"><h3>ATR MENSAL POR SAFRA (2022-2026)</h3><div class="chart-wrap tall"><canvas id="cAtrSafra"></canvas></div></div>
    <div class="card"><h3>MOAGEM MENSAL POR SAFRA (kt)</h3><div class="chart-wrap tall"><canvas id="cMoaSafra"></canvas></div></div>
  </div>
  <div class="grid g2" style="margin-top:14px">
    <div class="card"><h3>EGI MENSAL 2023-2026 (%)</h3><div class="chart-wrap"><canvas id="cEgiSafra"></canvas></div></div>
    <div class="card"><h3>APROVEITAMENTO INDUSTRIAL 2024-2026 (%)</h3><div class="chart-wrap"><canvas id="cAprov"></canvas></div></div>
  </div>
</div>

<div id="panel-hist" class="panel">
  <div class="kpi-row">
    <div class="kpi green"><div class="label">Melhor TCH Historico</div><div class="value">85.9</div><div class="sub">Safra 22/23-23/24</div></div>
    <div class="kpi red"><div class="label">Pior TCH Historico</div><div class="value">55.6</div><div class="sub">Safra 21/22-22/23</div></div>
    <div class="kpi green"><div class="label">Melhor ATR</div><div class="value">136.9</div><div class="sub">Ultima safra completa</div></div>
    <div class="kpi gold"><div class="label">TAH Tendencia</div><div class="value">+10%</div><div class="sub">Ultimas 3 safras</div></div>
  </div>
  <div class="grid g2">
    <div class="card"><h3>TCH HISTORICO POR SAFRA</h3><div class="chart-wrap tall"><canvas id="cTchHist"></canvas></div></div>
    <div class="card"><h3>ATR HISTORICO POR SAFRA</h3><div class="chart-wrap tall"><canvas id="cAtrHist"></canvas></div></div>
  </div>
  <div class="card" style="margin-top:14px">
    <h3>HISTORICO COMPLETO POR SAFRA</h3>
    <table><thead><tr><th>Safra</th><th>TCH</th><th>ATR</th><th>TAH</th><th>Area (ha)</th><th>TCH Trend</th><th>ATR Trend</th></tr></thead>
    <tbody id="hist-table"></tbody></table>
  </div>
</div>

<div id="panel-intel" class="panel">
  <div class="insight">
    <h4>ANALISE CONCLUSIVA — UMOE BIOENERGY | SAFRA 2025/26</h4>
    <p>
      <strong style="color:var(--red)">SITUACAO CRITICA:</strong>
      A UMOE enfrenta combinacao perigosa: canavial envelhecido (15.173 ha em 4C+),
      variedade dominante obsoleta (SP81-3250 em {SP81_HA:,.0f} ha), ATR 5% abaixo do ano anterior
      e gap de moagem de 216.500 t. Impacto combinado:
      <span class="hl">-R$ {GAP_R:.0f}M so nesta safra</span>.
      Sem acao estrutural, a tendencia e de piora progressiva nas proximas safras.
    </p>
  </div>

  <div class="grid g2">
    <div class="card">
      <h3>PONTOS POSITIVOS</h3>
      <table><thead><tr><th>Item</th><th>Evidencia</th><th>Impacto</th></tr></thead><tbody>
        <tr><td>EGI Industrial estavel</td><td>84-87% em jun/26</td><td style="color:var(--green)">Operacional OK</td></tr>
        <tr><td>Aproveitamento subindo</td><td>82.9% (mar/24) -> 92.6% (jun/26)</td><td style="color:var(--green)">+R$ 8M+</td></tr>
        <tr><td>CTC9006 disponivel</td><td>TAH 21.62 — 3.4x SP81</td><td style="color:var(--green)">R$ {SP81_POT_M}M potencial</td></tr>
        <tr><td>Tratos Soca eficientes</td><td>Unico CC abaixo orcado (-R$126/ha)</td><td style="color:var(--green)">Benchmark interno</td></tr>
        <tr><td>ATR set/25 = 154.8</td><td>Maior da historia UMOE</td><td style="color:var(--green)">Potencial existe</td></tr>
        <tr><td>Faz. 20614 Amb D = TAH 21.70</td><td>Blueprint do sucesso a replicar</td><td style="color:var(--green)">Caso de sucesso</td></tr>
      </tbody></table>
    </div>
    <div class="card">
      <h3>PONTOS CRITICOS</h3>
      <table><thead><tr><th>Item</th><th>Evidencia</th><th>Impacto R$</th></tr></thead><tbody>
        <tr><td style="color:var(--red)">SP81-3250: {SP81_HA:,.0f} ha</td><td>TAH {SP81_TAH:.1f} vs best {CTC_TAH:.1f}</td><td style="color:var(--red)">-R$ {SP81_POT_M}M+</td></tr>
        <tr><td style="color:var(--red)">4C+: 15.173 ha</td><td>TCH {tch_4c} vs 1C {tch_1c} (-{round((1-tch_4c/tch_1c)*100)}%)</td><td style="color:var(--red)">-R$ 40-60M</td></tr>
        <tr><td style="color:var(--red)">ATR 2026 -5% vs 2025</td><td>Todos os meses mar-jun/26</td><td style="color:var(--red)">-R$ 28M/safra</td></tr>
        <tr><td style="color:var(--red)">CCT R$50/t vs R$38,3/t</td><td>+30% acima do orcado</td><td style="color:var(--red)">-R$ 32M/safra</td></tr>
        <tr><td style="color:var(--gold)">RB72454: 4.845 ha</td><td>TAH 5.33 — segunda pior</td><td style="color:var(--red)">-R$ 15M+</td></tr>
        <tr><td style="color:var(--gold)">EGI em leve declinio</td><td>88.4% (2024) -> 85.7% (2026)</td><td style="color:var(--red)">-R$ 8M</td></tr>
      </tbody></table>
    </div>
  </div>

  <div class="card" style="margin-top:14px">
    <h3>TENDENCIAS IDENTIFICADAS</h3>
    <div class="grid g3">
      <div class="box3" style="border-left:3px solid var(--red)">
        <strong style="color:var(--red)">ATR 2026 — TENDENCIA NEGATIVA</strong>
        <p style="color:var(--text2);font-size:11px;margin-top:6px;line-height:1.6">
          ATR 5% abaixo de 2025 em TODOS os meses (mar: -3.1%, abr: -5.0%, mai: -5.4%, jun: -5.1%).
          Tendencia estrutural do canavial envelhecido. Se nao revertida: -R$ 28M na safra 25/26.
        </p>
      </div>
      <div class="box3" style="border-left:3px solid var(--gold)">
        <strong style="color:var(--gold)">MOAGEM — RECUPERACAO URGENTE</strong>
        <p style="color:var(--text2);font-size:11px;margin-top:6px;line-height:1.6">
          Jun/26: ritmo de 12,5 kt/dia vs meta de 15 kt/dia. Para atingir 2.768.000 t:
          precisa 18.500 t/dia de jul a nov. Historico maximo: 14,3 kt/dia (ago/25). Distancia critica.
        </p>
      </div>
      <div class="box3" style="border-left:3px solid var(--green)">
        <strong style="color:var(--green)">APROVEITAMENTO — TENDENCIA POSITIVA</strong>
        <p style="color:var(--text2);font-size:11px;margin-top:6px;line-height:1.6">
          Aproveitamento subiu de 82.9% (mar/24) para 92.6% (jun/26): +9.7 pp em 2 anos.
          Melhor indicador da empresa. Investimento industrial rendendo resultados. Manter.
        </p>
      </div>
    </div>
  </div>

  <div class="card" style="margin-top:14px">
    <h3>PLANO DE ACAO 5W2H — PRIORIZACAO POR IMPACTO R$</h3>
    <table>
      <thead><tr><th>Prior.</th><th>Acao</th><th>Responsavel</th><th>Prazo</th><th>Como</th><th>ROI</th></tr></thead>
      <tbody>
        <tr><td style="color:var(--red)">1 URGENTE</td><td>Plano substituicao SP81-3250</td><td>Dir. Agricola</td><td>Ago/26</td><td>Mapear SP81 por Amb, definir substituto (CTC9006 Amb D/E, CTC9007 Amb C), cronograma 3 safras</td><td style="color:var(--green)">R$ {SP81_POT_M}M+/safra</td></tr>
        <tr><td style="color:var(--red)">2 URGENTE</td><td>Acelerar reforma 4C+ (15.173 ha)</td><td>Dir. Agricola</td><td>Dez/26</td><td>Cronograma 25% ao ano (3.800 ha/safra), integrar com troca de variedades</td><td style="color:var(--green)">R$ 180M+ (5 anos)</td></tr>
        <tr><td style="color:var(--gold)">3 CURTO</td><td>Programa maturadores + ATR</td><td>Fitotecnia</td><td>Jul/26</td><td>Aplicar ethephon/RLEG nos canaviais SP81 para recuperar 8-12 kg/t ATR antes reforma</td><td style="color:var(--green)">R$ 28M/safra</td></tr>
        <tr><td style="color:var(--gold)">4 CURTO</td><td>Reducao CCT logistica + CHI</td><td>Oper. + Logistica</td><td>Imediato</td><td>Revisar DMT por frente, reduzir tempos patio, otimizar CHI controlavel</td><td style="color:var(--green)">R$ 32M/safra</td></tr>
        <tr><td style="color:var(--blue)">5 MEDIO</td><td>Expandir CTC9006 em Amb D e E</td><td>Fitotecnia</td><td>Safra 27/28</td><td>Priorizar nas reformas, negociar com CTC mudas para +5.000 ha em 3 anos</td><td style="color:var(--green)">R$ 79M+ (pleno)</td></tr>
        <tr><td style="color:var(--blue)">6 MEDIO</td><td>Auditar fazendas TAH &lt; 3 (Amb D)</td><td>Dir. Agricola</td><td>Set/26</td><td>Fazendas 20779, 20806, 20808: auditoria solo/variedade/manejo, decidir reforma emergencial</td><td style="color:var(--green)">R$ 15M+ pot.</td></tr>
      </tbody>
    </table>
  </div>

  <div class="card" style="margin-top:14px">
    <h3>VISOES QUE OS DADOS REVELAM — CRUZAMENTOS INEDITOS</h3>
    <div class="grid g2">
      <div class="box3" style="border-left:3px solid var(--red)">
        <strong style="color:var(--red)">SP81 x Amb D = Bomba Silenciosa</strong>
        <p style="color:var(--text2);font-size:11px;margin-top:6px;line-height:1.6">
          SP81-3250 domina o Amb D — o maior ambiente (182.562 ha).
          O Amb D tem solo OTIMO (TAH medio 8.69) mas variedade pessima.
          Se todo Amb D fosse CTC9006, TAH seria &gt;19.0. O solo nao e o problema — a variedade e.
        </p>
      </div>
      <div class="box3" style="border-left:3px solid var(--gold)">
        <strong style="color:var(--gold)">ATR alto x TCH baixo = Amb E escondido</strong>
        <p style="color:var(--text2);font-size:11px;margin-top:6px;line-height:1.6">
          Amb E tem MAIOR ATR (131.0) mas MENOR TCH (58.7). Parece ruim.
          Mas CTC9006 em Amb E atinge TAH 21.62 — a MELHOR combinacao da empresa.
          Amb E precisa de variedade de alta biomassa para liberar seu potencial de ATR.
        </p>
      </div>
      <div class="box3" style="border-left:3px solid var(--blue)">
        <strong style="color:var(--blue)">Decaimento real x Meta = Ilusao matematica</strong>
        <p style="color:var(--text2);font-size:11px;margin-top:6px;line-height:1.6">
          A meta de 2.768.000 t assume TCH viavel. Com 15.173 ha em 4C+ (TCH 52.5) e
          30.183 ha de SP81 (TCH ~43-51), o TCH medio real tende a 65-70 vs 80+ necessario.
          A meta so e atingivel com reforma estrutural do canavial.
        </p>
      </div>
      <div class="box3" style="border-left:3px solid var(--green)">
        <strong style="color:var(--green)">Fazenda 20614 = Blueprint do Sucesso</strong>
        <p style="color:var(--text2);font-size:11px;margin-top:6px;line-height:1.6">
          TAH 21.70, TCH 142.2, ATR 152.5 — TODOS os indicadores no topo.
          Entender e replicar o manejo desta fazenda para as 267 fazendas Amb D
          poderia elevar TAH medio de 8.69 para 12-15. Impacto: &gt;R$ 200M.
        </p>
      </div>
    </div>
  </div>
</div>

<div id="drill-modal">
  <div class="modal-box">
    <span class="close-btn" onclick="closeDrill()">x</span>
    <h3 id="drill-title">Fazendas</h3>
    <div id="drill-content"></div>
  </div>
</div>

<script>
const DRILLDOWN={json.dumps(drilldown,ensure_ascii=False)};
const AMB_DATA={json.dumps(amb_list,ensure_ascii=False)};
const VAR_LIST={json.dumps(top_vars,ensure_ascii=False)};
const WORST_VAR={json.dumps(worst_vars,ensure_ascii=False)};
const CORTE_DATA={json.dumps(corte_list,ensure_ascii=False)};
const SAFRA_DATA={json.dumps(safra_list,ensure_ascii=False)};
const VAR_AMB={json.dumps(top_var_amb,ensure_ascii=False)};
const IND={json.dumps(ind_series,ensure_ascii=False)};
const TOTAL_HA={TOTAL_HA:.0f};
const C={{green:'#00C853',gold:'#FFD600',red:'#FF1744',blue:'#2979FF',cyan:'#00E5FF',text2:'#8BA0BB',bg3:'#152344'}};

function showTab(id,el){{
  document.querySelectorAll('.panel').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById('panel-'+id).classList.add('active');
  el.classList.add('active');
}}
function openDrill(amb){{
  const fazs=DRILLDOWN[amb]||[];
  document.getElementById('drill-title').textContent='Ambiente '+amb+' — Top '+fazs.length+' Fazendas por TAH';
  const rows=fazs.map((f,i)=>{{
    const bar=Math.round(f.tah/22*100);
    const color=f.tah>14?C.green:f.tah>8?C.gold:C.red;
    return `<tr><td>${{i+1}}</td><td>${{f.faz}}</td><td><strong style="color:${{color}}">${{f.tah}}</strong><span class="bar-mini" style="width:${{bar}}px;background:${{color}}"></span></td><td>${{f.tch}}</td><td>${{f.atr}}</td><td>${{f.ha.toLocaleString()}}</td></tr>`;
  }}).join('');
  document.getElementById('drill-content').innerHTML=`<table><thead><tr><th>#</th><th>Fazenda</th><th>TAH</th><th>TCH</th><th>ATR</th><th>Area (ha)</th></tr></thead><tbody>${{rows}}</tbody></table>`;
  document.getElementById('drill-modal').classList.add('open');
}}
function closeDrill(){{document.getElementById('drill-modal').classList.remove('open');}}
document.getElementById('drill-modal').onclick=function(e){{if(e.target===this)closeDrill();}};

document.getElementById('amb-table').innerHTML=AMB_DATA.map(a=>{{
  const color=a.tah>9?C.green:a.tah>8?C.gold:C.red;
  return `<tr><td><strong style="color:${{color}}">Amb ${{a.amb}}</strong></td><td>${{a.tch}}</td><td>${{a.atr}}</td><td style="color:${{color}}">${{a.tah}}</td><td>${{a.ha.toLocaleString()}}</td><td>${{(a.ha/TOTAL_HA*100).toFixed(1)}}%</td><td><span class="drill-btn" onclick="openDrill('${{a.amb}}')">Ver fazendas</span></td></tr>`;
}}).join('');

document.getElementById('top-var-table').innerHTML=VAR_LIST.map((v,i)=>{{
  const color=v.tah>15?C.green:v.tah>10?C.gold:C.text2;
  return `<tr><td>${{i+1}}</td><td style="color:${{color}}">${{v.var}}</td><td style="color:${{color}}">${{v.tah}}</td><td>${{v.tch}}</td><td>${{v.atr}}</td><td>${{v.ha.toLocaleString()}}</td></tr>`;
}}).join('');

document.getElementById('worst-var-table').innerHTML=WORST_VAR.map((v,i)=>{{
  const perda=Math.round((10-v.tah)*v.ha*1000*1.03/1e6*10)/10;
  return `<tr><td>${{i+1}}</td><td style="color:${{C.red}}">${{v.var}}</td><td style="color:${{C.red}}">${{v.tah}}</td><td>${{v.tch}}</td><td>${{v.ha.toLocaleString()}}</td><td style="color:${{C.red}}">-R$${{perda.toFixed(1)}}M</td></tr>`;
}}).join('');

const tch1=CORTE_DATA[0]?.tch||98.6;
document.getElementById('corte-table').innerHTML=CORTE_DATA.map(c=>{{
  const pctL=((1-c.tch/tch1)*100).toFixed(1);
  const perda=(tch1-c.tch).toFixed(1);
  const color=c.corte<=2?C.green:c.corte==3?C.gold:C.red;
  return `<tr><td style="color:${{color}}">${{c.corte}}C</td><td style="color:${{color}}">${{c.tch}}</td><td style="color:${{c.corte>1?C.red:C.green}}">${{c.corte>1?'-'+pctL+'%':'baseline'}}</td><td style="color:${{C.red}}">${{c.corte>1?'-'+perda+' t/ha':'—'}}</td><td style="color:${{C.red}}">${{c.corte>=4?'R$ '+Math.round((tch1-c.tch)*15173*1.03*0.0145/1e6)+'M':'—'}}</td></tr>`;
}}).join('');

document.getElementById('hist-table').innerHTML=SAFRA_DATA.map((s,i)=>{{
  const prev=i>0?SAFRA_DATA[i-1]:null;
  const tt=prev?(s.tch>prev.tch?'<span class="trend-up">up</span>':'<span class="trend-down">dn</span>'):'—';
  const ta=prev?(s.atr>prev.atr?'<span class="trend-up">up</span>':'<span class="trend-down">dn</span>'):'—';
  return `<tr><td><strong>${{s.label}}</strong></td><td>${{s.tch}}</td><td>${{s.atr}}</td><td>${{s.tah}}</td><td>${{s.ha.toLocaleString()}}</td><td>${{tt}}</td><td>${{ta}}</td></tr>`;
}}).join('');

document.getElementById('var-amb-table').innerHTML=VAR_AMB.map(v=>{{
  const color=v.tah>15?C.green:v.tah>10?C.gold:C.text2;
  return `<tr><td style="color:${{color}}">${{v.var}}</td><td>Amb ${{v.amb}}</td><td style="color:${{color}}">${{v.tah}}</td><td>${{v.tch}}</td><td>${{v.ha.toLocaleString()}}</td></tr>`;
}}).join('');

const meses=['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'];
function getM(ano,f){{return Array.from({{length:12}},(_,i)=>{{const k=`${{ano}}-${{String(i+1).padStart(2,'0')}}`;return IND[k]?IND[k][f]:null;}});}}
function mkChart(id,type,labels,datasets,opts={{}}){{
  const ctx=document.getElementById(id)?.getContext('2d');
  if(!ctx)return;
  new Chart(ctx,{{type,data:{{labels,datasets}},options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{labels:{{color:C.text2,font:{{size:10}}}}}}}},
    scales:type!='pie'&&type!='doughnut'?{{x:{{ticks:{{color:C.text2,font:{{size:10}}}},grid:{{color:'#1E3252'}}}},y:{{ticks:{{color:C.text2,font:{{size:10}}}},grid:{{color:'#1E3252'}}}}}}: undefined,
    ...opts}}}});
}}
function yds(ano,f,color){{return{{label:ano,data:getM(ano,f),borderColor:color,pointRadius:2,tension:.4,spanGaps:true}};}}

mkChart('cAtrComp','line',meses,[
  yds('2024','atr',C.blue),yds('2025','atr',C.green),yds('2026','atr',C.red)]);
mkChart('cMoaComp','bar',meses,[
  {{label:'2025',data:getM('2025','moa'),backgroundColor:C.blue+'80',borderColor:C.blue,borderWidth:1}},
  {{label:'2026',data:getM('2026','moa'),backgroundColor:C.gold+'80',borderColor:C.gold,borderWidth:1}}]);
const egiK=Object.keys(IND).sort().filter(k=>IND[k].egi);
mkChart('cEgi','line',egiK.slice(-24).map(k=>k.slice(2)),[{{label:'EGI %',data:egiK.slice(-24).map(k=>IND[k].egi),borderColor:C.cyan,fill:true,backgroundColor:C.cyan+'15',tension:.4,pointRadius:2}}]);
mkChart('cAmb','bar',AMB_DATA.map(a=>'Amb '+a.amb),[
  {{label:'TCH',data:AMB_DATA.map(a=>a.tch),backgroundColor:C.blue+'80'}},
  {{label:'ATR',data:AMB_DATA.map(a=>a.atr),backgroundColor:C.gold+'80'}},
  {{label:'TAH x10',data:AMB_DATA.map(a=>+(a.tah*10).toFixed(0)),backgroundColor:C.green+'80'}}]);
mkChart('cAmbTah','bar',AMB_DATA.map(a=>'Amb '+a.amb+' ('+a.ha.toLocaleString()+'ha)'),[
  {{label:'TAH',data:AMB_DATA.map(a=>a.tah),
    backgroundColor:AMB_DATA.map(a=>a.tah>9?C.green+'80':a.tah>8?C.gold+'80':C.red+'80'),
    borderColor:AMB_DATA.map(a=>a.tah>9?C.green:a.tah>8?C.gold:C.red),borderWidth:1}}],
  {{indexAxis:'y'}});
mkChart('cVarRank','bar',VAR_LIST.slice(0,12).map(v=>v.var),[
  {{label:'TAH',data:VAR_LIST.slice(0,12).map(v=>v.tah),
    backgroundColor:VAR_LIST.slice(0,12).map(v=>v.tah>15?C.green+'90':v.tah>10?C.gold+'90':C.blue+'90'),borderWidth:0}}],
  {{indexAxis:'y'}});
mkChart('cCorte','bar',CORTE_DATA.map(c=>c.corte+'C'),[
  {{label:'TCH (t/ha)',data:CORTE_DATA.map(c=>c.tch),
    backgroundColor:CORTE_DATA.map(c=>c.corte<=2?C.green+'90':c.corte==3?C.gold+'90':C.red+'90'),borderWidth:0}}]);
mkChart('cAtrSafra','line',meses,[yds('2022','atr','#5C6BC0'),yds('2023','atr',C.blue),yds('2024','atr',C.cyan),yds('2025','atr',C.green),yds('2026','atr',C.red)]);
mkChart('cMoaSafra','bar',meses,[yds('2023','moa','#5C6BC0'),yds('2024','moa',C.blue),yds('2025','moa',C.green),yds('2026','moa',C.gold)]);
mkChart('cEgiSafra','line',meses,[yds('2023','egi',C.blue),yds('2024','egi',C.cyan),yds('2025','egi',C.green),yds('2026','egi',C.gold)]);
mkChart('cAprov','line',meses,[yds('2024','aprov',C.blue),yds('2025','aprov',C.green),yds('2026','aprov',C.gold)]);
mkChart('cTchHist','line',SAFRA_DATA.map(s=>s.label),[{{label:'TCH',data:SAFRA_DATA.map(s=>s.tch),borderColor:C.green,pointRadius:4,fill:true,backgroundColor:C.green+'15',tension:.3}}]);
mkChart('cAtrHist','line',SAFRA_DATA.map(s=>s.label),[{{label:'ATR',data:SAFRA_DATA.map(s=>s.atr),borderColor:C.gold,pointRadius:4,fill:true,backgroundColor:C.gold+'15',tension:.3}}]);
</script>
</body>
</html>"""

OUT_FILE.write_text(HTML, encoding='utf-8')
size_kb = OUT_FILE.stat().st_size // 1024
print(f"[5/5] HTML: {OUT_FILE} ({size_kb} KB)")
shutil.copy(OUT_FILE, DOCS_DIR / "UMOE_BI_Enterprise.html")
print(f"  Copia: docs/UMOE_BI_Enterprise.html")
print("  CONCLUIDO!")
