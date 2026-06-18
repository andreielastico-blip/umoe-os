# -*- coding: utf-8 -*-
"""
UMOE OS 8.0 | Cockpit Executivo da Safra 2026/27
Dashboard executivo profissional: KPIs vs meta, tendencias, comparativos,
cruzamentos e conclusoes concretas — tudo a partir dos dados reais extraidos
do Power BI (UMOE-OS-8.0/Dados-PBI/*.json).

Saida: UMOE-OS-8.0/Relatorios/UMOE_Cockpit_Executivo.html (+ copia em docs/)
"""
import json, sys
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd, numpy as np
except ImportError:
    import os; os.system("pip install pandas numpy -q"); import pandas as pd, numpy as np
import warnings; warnings.filterwarnings("ignore")

ROOT    = Path(__file__).parent.parent.parent
PBI_DIR = ROOT / "UMOE-OS-8.0" / "Dados-PBI"
OUT     = ROOT / "UMOE-OS-8.0" / "Relatorios" / "UMOE_Cockpit_Executivo.html"
DOCS    = ROOT / "docs"
HOJE    = datetime.now().strftime("%d/%m/%Y")

# ── Metas / parametros (SSoT UMOE 2026/27) ───────────────────────────────────
META_MOAGEM   = 2_768_000      # t safra 26/27
META_ATR      = 138.66         # kg/t
PRECO_ATR     = 1.03           # R$/kg CONSECANA

def load(name):
    """Carrega BASE_/CST_/CTRL_ JSON em DataFrame, limpando prefixo Tabela[Col]."""
    p = PBI_DIR / f"{name}.json"
    if not p.exists():
        return pd.DataFrame()
    try:
        d = json.load(open(p, encoding="utf-8-sig"))
    except Exception:
        return pd.DataFrame()
    if not d:
        return pd.DataFrame()
    df = pd.DataFrame(d)
    df.columns = [c.split("[")[-1].rstrip("]") if "[" in c else c for c in df.columns]
    return df

def load_manut(table):
    """Carrega tabela do workspace Manutencao (Dados-PBI/MANUT/umoe_dataset__<t>.json)."""
    p = PBI_DIR / "MANUT" / f"umoe_dataset__{table}.json"
    if not p.exists():
        return pd.DataFrame()
    try:
        d = json.load(open(p, encoding="utf-8-sig"))
    except Exception:
        return pd.DataFrame()
    if not d:
        return pd.DataFrame()
    df = pd.DataFrame(d)
    df.columns = [c.split("[")[-1].rstrip("]") if "[" in c else c for c in df.columns]
    return df

def num(s):
    return pd.to_numeric(s, errors="coerce").fillna(0)

def br(v, dec=0):
    """Formata numero no padrao BR (1.234,5)."""
    try:
        s = f"{float(v):,.{dec}f}"
        return s.replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "-"

print("[1/4] Carregando dados PBI...")
moa  = load("BASE_CTT_META_MOAGEM")
atrh = load("BASE_CTT_HISTPRD_ATR_HIST")
tch  = load("BASE_CTT_TCH")
tah  = load("BASE_CTT_TAH")
ind  = load("BASE_CTT_INDISPONIB")
cst  = load("CST_CST")
mat  = load("BASE_CTT_MATURADOR")
his  = load("BASE_CTT_HISTPRD")     # PRODUCAO REALIZADA (safra atual)
print(f"  realizado(HISTPRD)={len(his)} meta(MOAGEM)={len(moa)} tch={len(tch)} tah={len(tah)} indisp={len(ind)} custo={len(cst)}")

print("[2/4] Calculando metricas...")
K = {}   # KPIs
charts = {}
insights = []

# ── MOAGEM & ATR REALIZADOS (fonte: CTT_HISTPRD, safra 26/27) ───────────────
# Validado vs SSoT: mar-mai = 606.418 t e ATR ~126 kg/t.
moa_mensal = pd.DataFrame()
atr_mat = pd.DataFrame(); atr_var = pd.DataFrame(); atr_dia = pd.DataFrame(); atr_corte = pd.DataFrame()
if not his.empty:
    his["QT_CANA_ENT"] = num(his.get("QT_CANA_ENT", 0))
    his["KG_ACUCAR"]   = num(his.get("KG_ACUCAR", 0))
    his["DT"]          = pd.to_datetime(his.get("DT_HISTORICO"), errors="coerce")
    base = his[his["QT_CANA_ENT"] > 0].copy()
    tot_cana = base["QT_CANA_ENT"].sum()
    K["moagem_total"] = float(tot_cana)
    K["atend_meta"]   = tot_cana / META_MOAGEM * 100 if META_MOAGEM else 0
    K["atr_pond"]     = (base["KG_ACUCAR"].sum() / tot_cana) if tot_cana else 0
    K["dt_ini"] = str(base["DT"].min().date()) if base["DT"].notna().any() else ""
    K["dt_fim"] = str(base["DT"].max().date()) if base["DT"].notna().any() else ""

    # Tendencia mensal: moagem + ATR
    base["AM"] = base["DT"].dt.strftime("%Y-%m")
    g = base.dropna(subset=["DT"]).groupby("AM").agg(
        TON=("QT_CANA_ENT","sum"), AC=("KG_ACUCAR","sum")).reset_index()
    g["ATR"] = np.where(g["TON"]>0, g["AC"]/g["TON"], 0)
    moa_mensal = g.sort_values("AM")
    charts["moagem_mes"] = [{"MES": r["AM"], "TON": float(r["TON"]), "ATR": float(r["ATR"])} for _,r in moa_mensal.iterrows()]

    # ATR diario (todos os dias da safra)
    dd = base.dropna(subset=["DT"]).groupby(base["DT"].dt.date).apply(
        lambda x: x["KG_ACUCAR"].sum()/x["QT_CANA_ENT"].sum()).reset_index()
    dd.columns = ["dia","ATR"]; atr_dia = dd.sort_values("dia")
    charts["atr_dia"] = [{"dia": str(r["dia"]), "ATR": float(r["ATR"])} for _,r in atr_dia.iterrows()]

    def por(col):
        gg = base.groupby(col).apply(lambda x: pd.Series({
            "ATR": x["KG_ACUCAR"].sum()/x["QT_CANA_ENT"].sum(),
            "TON": x["QT_CANA_ENT"].sum()})).reset_index()
        return gg
    if "DE_MATURAC" in base.columns:
        atr_mat = por("DE_MATURAC").sort_values("ATR", ascending=False)
        charts["atr_mat"] = atr_mat.to_dict("records")
    if "DE_VARIED" in base.columns:
        av = por("DE_VARIED"); av = av[av["TON"] > tot_cana*0.005]
        atr_var = av.sort_values("ATR", ascending=False)
    if "ESTAGIO_CORTE" in base.columns:
        atr_corte = por("ESTAGIO_CORTE").sort_values("TON", ascending=False)

# ── META vs REALIZADO (META_MOAGEM.TON_EFET = meta oficial 2.768.542 t) ──────
meta_mes = pd.DataFrame()
if not moa.empty and not moa_mensal.empty:
    moa["TON_EFET"] = num(moa.get("TON_EFET", 0))
    moa["_DT"] = pd.to_datetime(moa.get("DT"), errors="coerce")
    moa["AM"] = moa["_DT"].dt.strftime("%Y-%m")
    gm = moa.groupby("AM")["TON_EFET"].sum().reset_index().rename(columns={"TON_EFET":"META"})
    gm = gm[gm["META"] > 0].sort_values("AM")
    real = moa_mensal[["AM","TON"]].rename(columns={"TON":"REAL"})
    mm = gm.merge(real, on="AM", how="left").fillna({"REAL": 0})
    mm["META_ACUM"] = mm["META"].cumsum()
    mm["REAL_ACUM"] = mm["REAL"].cumsum()
    meta_mes = mm
    charts["meta_mes"] = [{"AM": r["AM"], "META": float(r["META"]), "REAL": float(r["REAL"]),
                           "META_ACUM": float(r["META_ACUM"]), "REAL_ACUM": float(r["REAL_ACUM"])}
                          for _, r in mm.iterrows()]
    # Aderencia PONDERADA ate a data de analise (= ultima data com realizado).
    # Mes corrente parcial: a meta do mes e pro-rata pelos dias decorridos,
    # para meta x real serem comparaveis no mesmo numero de dias.
    import calendar
    adt = base["DT"].max() if "DT" in base.columns else None
    if adt is not None and pd.notna(adt):
        cur = adt.strftime("%Y-%m")
        dias_mes = calendar.monthrange(adt.year, adt.month)[1]
        frac = adt.day / dias_mes
        meta_ant = float(gm[gm["AM"] < cur]["META"].sum())
        meta_cur = float(gm[gm["AM"] == cur]["META"].sum()) * frac
        meta_td  = meta_ant + meta_cur
        K["data_analise"] = adt.strftime("%d/%m/%Y")
        K["frac_mes"] = frac
        K["_cur"] = cur
        # META de ATR ponderada ate a data (peso=TON_EFET; mes corrente pro-rata).
        # A meta de ATR sobe com a safra (mar 118 -> set 152); comparar com a
        # meta cheia (138,66) no inicio e injusto.
        moa["KG_ATR"] = num(moa.get("KG_ATR", 0))
        mw = moa[moa["TON_EFET"] > 0].copy()
        mw["w"] = mw["TON_EFET"] * mw["AM"].apply(lambda a: frac if a == cur else (1.0 if a < cur else 0.0))
        wsum = float(mw["w"].sum())
        K["meta_atr_periodo"] = float((mw["KG_ATR"]*mw["w"]).sum()/wsum) if wsum else META_ATR
        wfull = float(moa["TON_EFET"][moa["TON_EFET"]>0].sum())
        K["meta_atr_full"] = float((moa["KG_ATR"]*moa["TON_EFET"]).sum()/wfull) if wfull else META_ATR
        gatr = mw.groupby("AM").apply(lambda x:(x["KG_ATR"]*x["TON_EFET"]).sum()/x["TON_EFET"].sum() if x["TON_EFET"].sum() else 0).reset_index()
        gatr.columns=["AM","META_ATR"]
        charts["meta_atr_mes"] = [{"AM": r["AM"], "META_ATR": float(r["META_ATR"])} for _,r in gatr.iterrows()]
    else:
        meta_td = float(gm["META"].sum()); K["data_analise"] = ""
    real_td = float(K["moagem_total"])   # realizado total ate a data
    K["meta_total"]    = float(gm["META"].sum())
    K["meta_periodo"]  = meta_td
    K["real_periodo"]  = real_td
    K["atend_periodo"] = real_td/meta_td*100 if meta_td else 0
    K["gap_periodo"]   = real_td - meta_td

# ── TCH / TAH / VARIEDADES ──────────────────────────────────────────────────
tch_estagio = pd.DataFrame(); tah_var = pd.DataFrame()
if not tch.empty:
    tch["AREA_ESTIMADA"] = num(tch.get("AREA_ESTIMADA",0))
    tch["AREA_REEST3"]   = num(tch.get("AREA_REESTIMADA3",0))
    tch["PRD_REAL"] = num(tch.get("PRD_REAL",0))
    tch["PRD_ESTIMADA"] = num(tch.get("PRD_ESTIMADA",0))
    col_est = "ESTAGIO_AGRUP" if "ESTAGIO_AGRUP" in tch.columns else ("ESTAGIO_CORTE" if "ESTAGIO_CORTE" in tch.columns else None)
    if col_est:
        g = tch.groupby(col_est).agg(AREA=("AREA_ESTIMADA","sum"),
                                     AREA_R3=("AREA_REEST3","sum"),
                                     PRD_REAL=("PRD_REAL","sum"),
                                     PRD_EST=("PRD_ESTIMADA","sum")).reset_index()
        # FORMULA OFICIAL: TCH real = PRD_REAL/AREA_REESTIMADA3 ; TCH est = PRD_ESTIMADA/AREA_ESTIMADA
        g["TCH_REAL"] = np.where(g["AREA_R3"]>0, g["PRD_REAL"]/g["AREA_R3"], 0)
        g["TCH_EST"]  = np.where(g["AREA"]>0, g["PRD_EST"]/g["AREA"], 0)
        tch_estagio = g[(g["AREA"]>0)|(g["AREA_R3"]>0)].sort_values(col_est)
        tch_estagio = tch_estagio.rename(columns={col_est:"ESTAGIO"})
        charts["tch_estagio"] = tch_estagio.to_dict("records")
if not tah.empty:
    tah["AREA_CTT"]    = num(tah.get("AREA_CTT",0))
    tah["ACUCAR"]      = num(tah.get("ACUCAR",0))
    tah["TON"]         = num(tah.get("TON",0))
    tah["CANA_ACUCAR"] = num(tah.get("CANA_ACUCAR",0))
    if "DE_VARIED" in tah.columns:
        g = tah.groupby("DE_VARIED").agg(AREA=("AREA_CTT","sum"),
                                         ACUCAR=("ACUCAR","sum"),
                                         CANA_AC=("CANA_ACUCAR","sum"),
                                         TON=("TON","sum")).reset_index()
        # FORMULA OFICIAL Power BI (Z_CTT_TAH_TAH):
        # (ACUCAR/CANA_ACUCAR) * (TON/AREA) / 1000
        g["TAH"] = np.where((g["AREA"]>0)&(g["CANA_AC"]>0),
                            (g["ACUCAR"]/g["CANA_AC"])*(g["TON"]/g["AREA"])/1000.0, 0)
        g["TCH"] = np.where(g["AREA"]>0, g["TON"]/g["AREA"], 0)
        g = g[g["AREA"] > g["AREA"].sum()*0.01]
        tah_var = g.sort_values("TAH", ascending=False)

# ── CUSTOS ──────────────────────────────────────────────────────────────────
cst_grupo = pd.DataFrame()
if not cst.empty:
    cst["VL_ORC"]  = num(cst.get("VL_ORC",0))
    cst["VL_REAL"] = num(cst.get("VL_REAL",0))
    if "DT_REFER" in cst.columns:   # foca no ano corrente (2026)
        _dt = pd.to_datetime(cst["DT_REFER"], errors="coerce")
        cst = cst[_dt.dt.year == 2026]
    if "DE_GRUPO" in cst.columns and not cst.empty:
        g = cst.groupby("DE_GRUPO").agg(ORC=("VL_ORC","sum"), REAL=("VL_REAL","sum")).reset_index()
        g["DESVIO"] = g["REAL"] - g["ORC"]
        g["DESVIO_PCT"] = np.where(g["ORC"]>0, g["DESVIO"]/g["ORC"]*100, 0)
        cst_grupo = g.sort_values("REAL", ascending=False).head(12)
        K["custo_real"] = float(g["REAL"].sum())
        K["custo_orc"]  = float(g["ORC"].sum())

# ── CHUVA (fonte oficial: Excel Indice Pluviometrico, aba HISTORICO) ─────────
import glob as _glob
chuva_ano = pd.DataFrame(); chuva_mes = pd.DataFrame()
try:
    pl = r"C:\01 - UMOE\03 - Financeiro\Planilhas\1 - Indice Pluviometrico UMOE.xlsx"
    if Path(pl).exists():
        ch = pd.read_excel(pl, sheet_name="HISTORICO")
        ch.columns = [str(c).strip().upper() for c in ch.columns]
        qt = next((c for c in ch.columns if "LEITURA" in c), None)
        if qt and "ANO" in ch.columns:
            ch[qt] = num(ch[qt])
            dcol = next((c for c in ch.columns if c=="DATA"), None)
            ch["_M"] = pd.to_datetime(ch[dcol], errors="coerce").dt.month if dcol else 0
            # total anual (para o grafico de tendencia)
            ga = ch.groupby("ANO")[qt].sum().reset_index().rename(columns={qt:"MM"})
            ga = ga[(ga["ANO"]>=2015) & (ga["ANO"]<=2030)]
            chuva_ano = ga.sort_values("ANO")
            charts["chuva_ano"] = [{"ANO": int(r["ANO"]), "MM": float(r["MM"])} for _,r in chuva_ano.iterrows()]
            # YTD: mesmo periodo (ate o ultimo mes com dado em 2026) para comparacao justa
            mes_atual = int(ch[ch["ANO"]==2026]["_M"].max()) if (ch["ANO"]==2026).any() else 12
            ytd = ch[ch["_M"]<=mes_atual].groupby("ANO")[qt].sum()
            K["chuva_ano_atual"] = float(ytd.get(2026, 0))
            hist_ytd = ytd[(ytd.index>=2015) & (ytd.index<2026)]
            K["chuva_media"] = float(hist_ytd.mean()) if len(hist_ytd) else 0
            K["chuva_ytd_mes"] = mes_atual
            mcol = next((c for c in ch.columns if c in ("MÊS","MES")), None)
            if mcol:
                gm = ch[ch["ANO"]==2026].groupby(mcol)[qt].sum().reset_index().rename(columns={qt:"MM",mcol:"MES"})
                chuva_mes = gm
                charts["chuva_mes"] = [{"MES": str(r["MES"]), "MM": float(r["MM"])} for _,r in gm.iterrows()]
except Exception as e:
    print("  chuva: erro", e)

# ── VARIEDADES (historico BD SAFRAS, 8 safras) ───────────────────────────────
var_hist = pd.DataFrame(); var_amb = pd.DataFrame(); hist_safra = pd.DataFrame()
try:
    bds = _glob.glob(r"C:\01 - UMOE\03 - Financeiro\Planilhas\Historico Kpis*Agr*.xlsx")
    if bds:
        wb = pd.ExcelFile(bds[0])
        sh = next((s for s in wb.sheet_names if "BD" in s.upper() and "SAFRA" in s.upper()), wb.sheet_names[0])
        bd = wb.parse(sh, header=0)
        # mapeamento por posicao (umoe-bi-enterprise): 0=Amb,1=Faz,5=Var,7=Area,11=TCH,16=ATR,17=TAH,21=Safra,24=Corte
        cols = list(bd.columns)
        def col(i): return cols[i] if i < len(cols) else None
        bd = bd.rename(columns={col(0):"AMB",col(1):"FAZ",col(5):"VAR",col(7):"AREA",
                                col(11):"TCH",col(16):"ATR",col(17):"TAH",col(21):"SAFRA",col(24):"CORTE"})
        for c in ["AREA","TCH","ATR","TAH"]:
            if c in bd.columns: bd[c] = num(bd[c])
        bd = bd[(bd.get("AREA",0)>0.5) & (bd.get("TCH",0)>10) & (bd.get("TAH",0)>0.5)]
        if "VAR" in bd.columns and not bd.empty:
            g = bd.groupby("VAR").apply(lambda x: pd.Series({
                "TAH": (x["TAH"]*x["AREA"]).sum()/x["AREA"].sum(),
                "TCH": (x["TCH"]*x["AREA"]).sum()/x["AREA"].sum(),
                "AREA": x["AREA"].sum()})).reset_index()
            g = g[g["AREA"] > g["AREA"].sum()*0.005]
            var_hist = g.sort_values("TAH", ascending=False)
            charts["var_hist"] = var_hist.head(15).to_dict("records")
        if "VAR" in bd.columns and "AMB" in bd.columns and not bd.empty:
            bd["AMB1"] = bd["AMB"].astype(str).str[:1]
            ga = bd[bd["AMB1"].str.len()==1].groupby(["VAR","AMB1"]).apply(
                lambda x: pd.Series({"TAH": (x["TAH"]*x["AREA"]).sum()/x["AREA"].sum(),
                                     "AREA": x["AREA"].sum()})).reset_index()
            var_amb = ga[ga["AREA"]>300].sort_values("TAH", ascending=False)
        K["var_safras"] = int(bd["SAFRA"].nunique()) if "SAFRA" in bd.columns else 0
        K["var_registros"] = int(len(bd))
        # Historico por safra (TCH/ATR/TAH ponderado por area)
        if "SAFRA" in bd.columns:
            gs = bd.groupby("SAFRA").apply(lambda x: pd.Series({
                "TCH": (x["TCH"]*x["AREA"]).sum()/x["AREA"].sum(),
                "ATR": (x["ATR"]*x["AREA"]).sum()/x["AREA"].sum(),
                "TAH": (x["TAH"]*x["AREA"]).sum()/x["AREA"].sum()})).reset_index()
            def saf(c):  # 21213 -> 12/13
                s=str(int(c));
                return s[1:3]+"/"+s[3:5] if len(s)>=5 else s
            gs["SAFRA_LBL"] = gs["SAFRA"].apply(saf)
            gs = gs.sort_values("SAFRA")
            hist_safra = gs
            charts["hist_safra"] = [{"safra": r["SAFRA_LBL"], "TCH": float(r["TCH"]),
                                     "ATR": float(r["ATR"]), "TAH": float(r["TAH"])} for _,r in gs.iterrows()]
            # adiciona a safra atual (26/27) com ATR realizado e TCH/TAH atuais
            cur = {"safra":"26/27*", "ATR": float(K.get("atr_pond",0)),
                   "TCH": float(tah_var["TCH"].mean()) if not tah_var.empty else 0,
                   "TAH": float(tah_var["TAH"].mean()) if not tah_var.empty else 0}
            charts["hist_safra"].append(cur)
except Exception as e:
    print("  bd_safras: erro", e)

# ── COMPARATIVO ENTRE SAFRAS POR MES (historico industrial diario) ───────────
# Fonte: Histórico Industrias Diário Safras.xlsx (1 aba por ano, col2=data,
# col3=moagem dia, col4=ATR dia). Compara safras no MESMO ponto da safra.
moagem_pace = pd.DataFrame()
try:
    import openpyxl
    fI = _glob.glob(r"C:\01 - UMOE\03 - Financeiro\Planilhas\Hist*Industri*Safras.xlsx")
    if fI:
        wbI = openpyxl.load_workbook(fI[0], read_only=True, data_only=True)
        anos = [a for a in ["2022","2023","2024","2025","2026"] if a in wbI.sheetnames]
        regs = []
        for ano in anos:
            for r in wbI[ano].iter_rows(min_row=3, values_only=True):
                dt = r[2] if len(r)>2 else None
                if dt is None or not hasattr(dt, "month"):
                    continue
                mo = float(r[3] or 0) if len(r)>3 and r[3] is not None else 0
                at = float(r[4] or 0) if len(r)>4 and r[4] is not None else 0
                regs.append((int(ano), dt.month, dt.day, mo, at))
        di = pd.DataFrame(regs, columns=["ANO","MES","DIA","MOA","ATR"])
        # cutoff = ultima data com moagem na safra corrente (2026)
        cur = di[(di["ANO"]==2026) & (di["MOA"]>0)]
        if not cur.empty:
            cmes = int(cur["MES"].max()); cdia = int(cur[cur["MES"]==cmes]["DIA"].max())
        else:
            cmes, cdia = 12, 31
        ate = di[(di["MES"]<cmes) | ((di["MES"]==cmes) & (di["DIA"]<=cdia))]
        # ATR mensal ponderado por moagem, por safra
        gm2 = di[di["MOA"]>0].copy(); gm2["AC"]=gm2["MOA"]*gm2["ATR"]
        am = gm2.groupby(["ANO","MES"]).agg(MOA=("MOA","sum"),AC=("AC","sum")).reset_index()
        am["ATR"]=am["AC"]/am["MOA"]
        meses = sorted(am["MES"].unique())
        charts["atr_mes_safra"] = {"meses": [int(m) for m in meses],
            "series": {str(a): [float(am[(am["ANO"]==a)&(am["MES"]==m)]["ATR"].sum() or 0) for m in meses] for a in anos}}
        # Moagem acumulada por safra ate a mesma data (ritmo) + ATR ate a data
        gp = ate[ate["MOA"]>0].groupby("ANO").agg(MOA=("MOA","sum")).reset_index()
        gpa = ate.copy(); gpa["AC"]=gpa["MOA"]*gpa["ATR"]
        atr_ate = gpa.groupby("ANO").apply(lambda x: x["AC"].sum()/x["MOA"].sum() if x["MOA"].sum() else 0)
        gp["ATR"]=gp["ANO"].map(atr_ate)
        moagem_pace = gp.sort_values("ANO")
        charts["moagem_pace"] = [{"safra": f"{a-1 if False else str(a)[2:]}/{str(a+1)[2:]}",
                                  "MOA": float(moagem_pace[moagem_pace["ANO"]==a]["MOA"].sum()),
                                  "ATR": float(moagem_pace[moagem_pace["ANO"]==a]["ATR"].sum())} for a in moagem_pace["ANO"]]
        K["pace_cutoff"] = f"{cdia:02d}/{cmes:02d}"
except Exception as e:
    print("  industrial: erro", e)

# ── LOGISTICA / CARGAS (cana queimada, KM medio) ─────────────────────────────
carga_kpi = {}
try:
    cg = load("BASE_CTT_CARGAS_ENTRADA")
    if not cg.empty:
        cg["QT_CARGA"] = num(cg.get("QT_CARGA",0)); cg["KM_POND"] = num(cg.get("KM_POND",0)); cg["TON_ESTIM"]=num(cg.get("TON_ESTIM",0))
        tot = cg["TON_ESTIM"].sum()
        if "NO_QUEIMA" in cg.columns and tot:
            q = cg.copy(); q["queim"] = q["NO_QUEIMA"].astype(str).str.lower().str.contains("queim")
            carga_kpi["perc_queima"] = float(q[q["queim"]]["TON_ESTIM"].sum()/tot*100)
        if tot:
            carga_kpi["km_medio"] = float((cg["KM_POND"]*cg["TON_ESTIM"]).sum()/tot)
        carga_kpi["cargas"] = int(len(cg))
except Exception as e:
    print("  cargas: erro", e)

# ── PARADAS / INDISPONIBILIDADE ─────────────────────────────────────────────
paradas = pd.DataFrame()
if not ind.empty:
    ind["HORAS_DEC"] = num(ind.get("HORAS_DEC",0))
    col_g = "GRUPO_PARADA" if "GRUPO_PARADA" in ind.columns else ("AGRP_PARADA" if "AGRP_PARADA" in ind.columns else None)
    if col_g:
        g = ind.groupby(col_g)["HORAS_DEC"].sum().reset_index().rename(columns={col_g:"GRUPO","HORAS_DEC":"HORAS"})
        paradas = g[g["HORAS"]>0].sort_values("HORAS", ascending=False).head(10)
        charts["paradas"] = paradas.to_dict("records")

# ── FRENTES: PROPRIA x LEROSA x FABIANO (moagem + ATR) ───────────────────────
frentes = pd.DataFrame()
GRP_FRENTE = {1:"Propria",2:"Propria",3:"Propria",4:"Propria",10:"Lerosa (F10)",27:"Fabiano (F27)"}
if not his.empty and "CD_FREN_TRAN" in his.columns:
    hf = his.copy(); hf["CD_FREN_TRAN"]=pd.to_numeric(hf["CD_FREN_TRAN"],errors="coerce")
    hf["GRP"]=hf["CD_FREN_TRAN"].map(GRP_FRENTE).fillna("Outras")
    g = hf.groupby("GRP").apply(lambda x: pd.Series({
        "CANA": x["QT_CANA_ENT"].sum(),
        "ATR": x["KG_ACUCAR"].sum()/x["QT_CANA_ENT"].sum() if x["QT_CANA_ENT"].sum() else 0})).reset_index()
    frentes = g[g["CANA"]>0].sort_values("CANA", ascending=False)
    # META OFICIAL de moagem por frente (CTT_META_MOAGEM.TON_EFET), pro-rata a data
    if not moa.empty and "CD_FREN_TRAN" in moa.columns and "frac_mes" in K:
        mo = moa.copy(); mo["TON_EFET"]=num(mo.get("TON_EFET",0))
        mo["GRP"]=pd.to_numeric(mo["CD_FREN_TRAN"],errors="coerce").map(GRP_FRENTE).fillna("Outras")
        cur=K.get("_cur"); fr=K.get("frac_mes",1.0)
        mo["w"]=mo["AM"].apply(lambda a: fr if a==cur else (1.0 if (cur and a<cur) else 0.0))
        mg=mo.groupby("GRP").apply(lambda x:(x["TON_EFET"]*x["w"]).sum()).reset_index()
        mg.columns=["GRP","META"]
        frentes=frentes.merge(mg,on="GRP",how="left").fillna({"META":0})
        frentes["ADER"]=np.where(frentes["META"]>0, frentes["CANA"]/frentes["META"]*100, 0)
    else:
        frentes["META"]=0; frentes["ADER"]=0
    charts["frentes"] = frentes.to_dict("records")

# ── PROJECAO DE FIM DE SAFRA (ritmo atual vs meta cheia) ─────────────────────
if K.get("meta_periodo") and K.get("meta_total"):
    pace = K["real_periodo"]/K["meta_periodo"] if K["meta_periodo"] else 0
    K["proj_final"] = pace * K["meta_total"]
    K["proj_gap"]   = K["proj_final"] - K["meta_total"]
    K["pace_pct"]   = pace*100

# ── ATR FULL POR SAFRA (industrial, inclui 24/25 e 25/26) ────────────────────
# (di e calculado no bloco COMPARATIVO acima)
try:
    if "di" in dir() and not di.empty:
        gf = di[di["MOA"]>0].copy(); gf["AC"]=gf["MOA"]*gf["ATR"]
        gy = gf.groupby("ANO").apply(lambda x: pd.Series({
            "ATR": x["AC"].sum()/x["MOA"].sum(), "MOA": x["MOA"].sum()})).reset_index()
        gy["LBL"]=gy["ANO"].apply(lambda a: f"{str(a)[2:]}/{str(a+1)[2:]}")
        charts["atr_safra_full"]=[{"safra":r["LBL"],"ATR":float(r["ATR"]),"MOA":float(r["MOA"])} for _,r in gy.sort_values("ANO").iterrows()]
except Exception as e: print("  atr_safra_full:", e)

# ── PRAGAS (broca, cigarrinha) ───────────────────────────────────────────────
praga = {}
try:
    brc = load("BASE_CTT_BROCA")
    if not brc.empty:
        brc["KG_CANA_ENT"]=num(brc.KG_CANA_ENT); brc["KG_CANA_BROCA"]=num(brc.KG_CANA_BROCA)
        praga["broca"]= float(brc["KG_CANA_BROCA"].sum()/brc["KG_CANA_ENT"].sum()*100) if brc["KG_CANA_ENT"].sum() else 0
        bf = brc.groupby("FAZENDA").apply(lambda x: x["KG_CANA_BROCA"].sum()/x["KG_CANA_ENT"].sum()*100 if x["KG_CANA_ENT"].sum() else 0).reset_index()
        bf.columns=["FAZENDA","INFEST"]; praga["broca_top"]=bf.sort_values("INFEST",ascending=False).head(10).to_dict("records")
        charts["broca_top"]=praga["broca_top"]
    cg = load("BASE_AGRO_CIGARRINHA")
    if not cg.empty:
        for c in ["QT_ADULTAS","QT_NINFAS","QT_AMOSTRA","QT_AREA_PROD"]: cg[c]=num(cg.get(c,0))
        praga["cig_total"]= float(cg["QT_ADULTAS"].sum()+cg["QT_NINFAS"].sum())
        praga["cig_fazendas"]= int(cg["FAZENDA"].nunique()) if "FAZENDA" in cg.columns else 0
except Exception as e: print("  pragas:", e)

# ── PLANTIO (real vs meta) ───────────────────────────────────────────────────
plantio = {}
try:
    pr = load("BASE_PLANT_REAL"); pm = load("BASE_PLANT_META")
    if not pr.empty:
        plantio["real_ha"]= float(num(pr.get("QT_AREA",0)).sum())
    if not pm.empty and "HA" in pm.columns:
        plantio["meta_ha"]= float(num(pm["HA"]).sum())
    if plantio.get("meta_ha"):
        plantio["aderencia"]= plantio.get("real_ha",0)/plantio["meta_ha"]*100
except Exception as e: print("  plantio:", e)

# ── ADERENCIA DE TRATOS (ADEREN_ATIVIDADES: area realizada vs prevista) ───────
ader_proc = pd.DataFrame()
try:
    ad = load("BASE_ADEREN_ATIVIDADES")
    if not ad.empty:
        ad["AREA"]=num(ad.get("AREA",0)); ad["AREA_REALIZADA"]=num(ad.get("AREA_REALIZADA",0))
        col = "DESC_PROCESSO" if "DESC_PROCESSO" in ad.columns else ("PROCESSO" if "PROCESSO" in ad.columns else None)
        if col:
            g = ad.groupby(col).agg(PREV=("AREA","sum"), REAL=("AREA_REALIZADA","sum")).reset_index().rename(columns={col:"PROC"})
            g["ADER"] = np.where(g["PREV"]>0, g["REAL"]/g["PREV"]*100, 0)
            ader_proc = g[g["PREV"]>0].sort_values("PREV", ascending=False).head(14)
            charts["ader_proc"]=[{"proc":str(r["PROC"]),"prev":float(r["PREV"]),"real":float(r["REAL"]),"ader":float(r["ADER"])} for _,r in ader_proc.iterrows()]
        tot_prev=ad["AREA"].sum()
        K["ader_geral"]= float(ad["AREA_REALIZADA"].sum()/tot_prev*100) if tot_prev else 0
        K["ader_prev_ha"]=float(tot_prev); K["ader_real_ha"]=float(ad["AREA_REALIZADA"].sum())
except Exception as e: print("  aderencia:", e)

# ── MANUTENCAO / FROTA (workspace Manutencao Premium) ────────────────────────
manut = {}; diesel_cat = pd.DataFrame(); meta_disp = pd.DataFrame()
try:
    # Diesel (f_abastecimento) - foca 2026
    ab = load_manut("f_abastecimento")
    if not ab.empty:
        ab["LITROS"]=num(ab.get("LITROS",0)); ab["_DT"]=pd.to_datetime(ab.get("DT_OPERACAO"),errors="coerce")
        a26 = ab[ab["_DT"].dt.year==2026]
        base_ab = a26 if not a26.empty else ab
        manut["diesel_l"]=float(base_ab["LITROS"].sum())
        dc = base_ab.groupby("CATEGORIA")["LITROS"].sum().reset_index().sort_values("LITROS",ascending=False).head(8)
        diesel_cat = dc; charts["diesel_cat"]=[{"cat":str(r["CATEGORIA"]),"L":float(r["LITROS"])} for _,r in dc.iterrows()]
    # Produtividade transporte (DMT) - 2026
    fp = load_manut("f_produtividade")
    if not fp.empty:
        for c in ["QT_CANA_ENT","QT_CARGA_ENT","QT_VIAG_ENT","DISTANCIA"]: fp[c]=num(fp.get(c,0))
        fp["_DT"]=pd.to_datetime(fp.get("DT_HISTORICO"),errors="coerce")
        p26 = fp[fp["_DT"].dt.year==2026]; bp = p26 if not p26.empty else fp
        vsum = bp["QT_VIAG_ENT"].sum()
        manut["dmt"]=float((bp["DISTANCIA"]*bp["QT_VIAG_ENT"]).sum()/vsum) if vsum else float(bp["DISTANCIA"].mean())
        manut["viagens"]=float(bp["QT_VIAG_ENT"].sum()); manut["cargas"]=float(bp["QT_CARGA_ENT"].sum())
    # Manutencao (f_manfro) - corretiva vs preventiva, 2026
    mf = load_manut("f_manfro")
    if not mf.empty:
        mf["_DT"]=pd.to_datetime(mf.get("ENTRADA"),errors="coerce")
        m26 = mf[mf["_DT"].dt.year==2026]; bm = m26 if not m26.empty else mf
        manut["os"]=int(len(bm))
        cl = bm["CLASSE_MNT"].astype(str)
        corr = int(cl.str.contains("orretiv",case=False).sum())
        manut["corretiva_pct"]=corr/len(bm)*100 if len(bm) else 0
        manut["horas_parado"]=float(num(bm.get("TEMPO_TOTAL_PARADO",0)).sum()) if "TEMPO_TOTAL_PARADO" in bm.columns else 0
        # top equipamentos por OS
        te = bm.groupby("CD_EQUIPTO").size().reset_index(name="OS").sort_values("OS",ascending=False).head(10)
        charts["manut_top"]=[{"eq":str(r["CD_EQUIPTO"]),"OS":int(r["OS"])} for _,r in te.iterrows()]
    # Metas de disponibilidade (alvos oficiais)
    md = load_manut("d_meta_equipamentos")
    if not md.empty:
        mdcol = next((c for c in md.columns if "META" in c.upper() and "DISP" in c.upper()), None)
        if mdcol:
            md[mdcol]=num(md[mdcol])
            meta_disp = md[md[mdcol]>0][["CATEGORIA",mdcol]].rename(columns={mdcol:"META_DISP"}).sort_values("META_DISP",ascending=False)
    # DISPONIBILIDADE REAL OFICIAL (medida DAX, consultada e salva por refresh; mes corrente)
    dfp = PBI_DIR / "MANUT" / "disponibilidade_oficial.json"
    if dfp.exists():
        do = json.load(open(dfp, encoding="utf-8"))
        g = do.get("geral", {}) if isinstance(do, dict) else {}
        if g.get("disp") is not None:
            manut["disp_real"]=float(g["disp"]); manut["disp_meta"]=float(g.get("meta") or 0)
        charts["disp_mensal"]=[{"ym": r["ym"], "disp": float(r["disp"]), "meta": float(r["meta"])} for r in do.get("mensal", [])]
        cats=sorted(do.get("categoria", []), key=lambda r: r.get("disp",100))[:12]
        charts["disp_cat"]=[{"cat": str(r["cat"]), "disp": float(r["disp"]), "meta": float(r.get("meta") or 0)} for r in cats]
except Exception as e:
    print("  manutencao:", e)

# ── FINANCEIRO: receita ATR (CONSECANA) + gap projetado ──────────────────────
if K.get("atr_pond"):
    K["receita_real"] = K["moagem_total"] * K["atr_pond"] * PRECO_ATR          # R$ (kg ATR x R$/kg)
    if K.get("proj_final"):
        K["receita_proj"] = K["proj_final"] * K.get("meta_atr_full",META_ATR) * PRECO_ATR
        K["receita_meta"] = K.get("meta_total",META_MOAGEM) * K.get("meta_atr_full",META_ATR) * PRECO_ATR
        K["receita_gap"]  = K["receita_proj"] - K["receita_meta"]

# ── PERDA INDUSTRIAL (CTT_META_MOAGEM: TON_PERDA real vs PERC_PERDA meta) ─────
if not moa.empty:
    moa["TON"] = num(moa.get("TON",0)); moa["TON_PERDA"]=num(moa.get("TON_PERDA",0)); moa["PERC_PERDA"]=num(moa.get("PERC_PERDA",0))
    tton = moa["TON"].sum()
    if tton:
        K["perda_real_pct"] = float(moa["TON_PERDA"].sum()/tton*100)
        K["perda_meta_pct"] = float((moa["PERC_PERDA"]*moa["TON"]).sum()/tton*100)

# ── INSIGHTS (conclusoes concretas) ─────────────────────────────────────────
def add(txt, tipo="info"): insights.append({"t": txt, "tipo": tipo})

if "moagem_total" in K:
    if "meta_periodo" in K:
        sit = "verde" if K["atend_periodo"]>=98 else ("amarelo" if K["atend_periodo"]>=90 else "vermelho")
        add(f"Aderencia a meta ate {K.get('data_analise','')} (meta pro-rata aos dias decorridos): realizado {br(K['real_periodo'])} t vs "
            f"meta {br(K['meta_periodo'])} t = {br(K['atend_periodo'],1)}% do ritmo. {'Atraso' if K['gap_periodo']<0 else 'Adiantado'} de "
            f"{br(abs(K['gap_periodo']))} t. Meta cheia da safra: {br(K.get('meta_total',META_MOAGEM))} t.", sit)
    else:
        add(f"Moagem acumulada de {br(K['moagem_total'])} t = {br(K['atend_meta'],1)}% da meta safra ({br(META_MOAGEM)} t).", "amarelo")
if "atr_pond" in K and K["atr_pond"]>0:
    mp = K.get("meta_atr_periodo", META_ATR)
    d = K["atr_pond"] - mp
    add(f"ATR realizado {br(K['atr_pond'],2)} kg/t vs META PONDERADA ate a data {br(mp,2)} kg/t "
        f"({'+' if d>=0 else ''}{br(d,2)} kg/t) — comparacao justa (a meta sobe com a maturacao). "
        f"Meta da safra cheia: {br(K.get('meta_atr_full',META_ATR),2)} kg/t, atingida com o ATR alto de set-out. "
        f"Cada kg/t vale ~R$ {br(K.get('moagem_total',0)*PRECO_ATR/1000)} no acumulado a CONSECANA.",
        "verde" if d>=-1.5 else ("amarelo" if d>=-4 else "vermelho"))
if not atr_mat.empty:
    top = atr_mat.iloc[0]
    add(f"Maturacao '{top['DE_MATURAC']}' concentra o maior ATR ({br(top['ATR'],1)} kg/t). "
        f"Priorizar colheita por janela de maturacao maximiza acucar.", "info")
if not tah_var.empty:
    b = tah_var.iloc[0]; w = tah_var.iloc[-1]
    add(f"Variedade lider em TAH: {b['DE_VARIED']} ({br(b['TAH'],2)} t ATR/ha). "
        f"Pior relevante: {w['DE_VARIED']} ({br(w['TAH'],2)}). Diferenca de {br(b['TAH']-w['TAH'],2)} t ATR/ha entre extremos.", "info")
if not cst_grupo.empty:
    est = cst_grupo[cst_grupo["DESVIO"]>0].sort_values("DESVIO", ascending=False)
    if not est.empty:
        e = est.iloc[0]
        add(f"Maior estouro de custo: grupo '{e['DE_GRUPO']}' com R$ {br(e['DESVIO'])} acima do orcado "
            f"({'+' if e['DESVIO_PCT']>=0 else ''}{br(e['DESVIO_PCT'],1)}%).", "vermelho")
if not paradas.empty:
    p = paradas.iloc[0]
    add(f"Maior ofensor de paradas: '{p['GRUPO']}' com {br(p['HORAS'],1)} h acumuladas. Foco de ganho de disponibilidade.", "amarelo")
if "chuva_ano_atual" in K and K.get("chuva_media",0)>0:
    d = (K["chuva_ano_atual"]/K["chuva_media"]-1)*100
    mm = K.get("chuva_ytd_mes",12)
    add(f"Chuva 2026 (jan-mes {mm}, YTD): {br(K['chuva_ano_atual'])} mm vs media historica do mesmo periodo {br(K['chuva_media'])} mm "
        f"({'+' if d>=0 else ''}{br(d,0)}%). Clima e o maior ofensor de paradas — correlacao direta com disponibilidade de moagem.",
        "info")
if not var_hist.empty:
    b = var_hist.iloc[0]; w = var_hist.iloc[-1]
    add(f"Historico (8 safras): {b['VAR']} lidera TAH ({br(b['TAH'],2)} t ATR/ha); {w['VAR']} e a pior relevante "
        f"({br(w['TAH'],2)}). Direcionar plantio para as variedades-topo eleva o ATR estrutural.", "info")
if not var_amb.empty:
    top = var_amb.iloc[0]
    add(f"Melhor combinacao variedade x ambiente: {top['VAR']} x Amb.{top['AMB1']} = {br(top['TAH'],2)} t ATR/ha. "
        f"Alocacao por aptidao de ambiente maximiza produtividade.", "verde")
if not moagem_pace.empty and (moagem_pace["ANO"]==2026).any():
    m26 = float(moagem_pace[moagem_pace["ANO"]==2026]["MOA"].sum())
    ant = moagem_pace[moagem_pace["ANO"]<2026]
    if not ant.empty:
        med = float(ant["MOA"].mean())
        d = (m26/med-1)*100 if med else 0
        a26 = float(moagem_pace[moagem_pace["ANO"]==2026]["ATR"].sum())
        amed = float(ant["ATR"].mean())
        add(f"Ritmo vs historico (mesmo ponto, ate {K.get('pace_cutoff','')}): moagem 26/27 {br(m26)} t = "
            f"{'+' if d>=0 else ''}{br(d,0)}% vs media das ultimas safras ({br(med)} t). "
            f"ATR 26/27 {br(a26,1)} kg/t vs media {br(amed,1)} kg/t.", "verde" if d>=0 else "vermelho")

# Insights operacionais adicionais
if not frentes.empty and "ADER" in frentes.columns and (frentes["META"]>0).any():
    fa = frentes[frentes["META"]>0].sort_values("ADER")
    pior = fa.iloc[0]
    add(f"Aderencia a meta por frente (realizado vs meta oficial pro-rata): pior e {pior['GRP']} com {br(pior['ADER'],0)}% "
        f"({br(pior['CANA'])} t vs meta {br(pior['META'])} t). Melhor: {fa.iloc[-1]['GRP']} ({br(fa.iloc[-1]['ADER'],0)}%).",
        "vermelho" if pior['ADER']<85 else "amarelo")
if not frentes.empty:
    fr = frentes.set_index("GRP")
    if "Fabiano (F27)" in fr.index and "Lerosa (F10)" in fr.index:
        add(f"Qualidade por frente: Lerosa (F10) entrega ATR {br(fr.loc['Lerosa (F10)','ATR'],1)} kg/t, "
            f"Propria {br(fr.loc['Propria','ATR'],1) if 'Propria' in fr.index else '-'}, "
            f"Fabiano (F27) so {br(fr.loc['Fabiano (F27)','ATR'],1)} kg/t — diferenca de "
            f"{br(fr.loc['Lerosa (F10)','ATR']-fr.loc['Fabiano (F27)','ATR'],1)} kg/t. Cobrar maturacao/ponto de corte do fornecedor Fabiano.", "vermelho")
if K.get("receita_gap") is not None:
    add(f"FINANCEIRO (CONSECANA R$ {br(PRECO_ATR,2)}/kg ATR): receita realizada ~R$ {br(K['receita_real']/1e6,1)} M. "
        f"No ritmo atual, projecao de receita ~R$ {br(K['receita_proj']/1e6,0)} M vs meta R$ {br(K['receita_meta']/1e6,0)} M "
        f"= risco de {br(K['receita_gap']/1e6,0)} M (deficit de moagem se nao recuperado).",
        "vermelho" if K['receita_gap']<0 else "verde")
if K.get("perda_real_pct") is not None:
    dpf = K["perda_real_pct"]-K.get("perda_meta_pct",0)
    add(f"Perda industrial: {br(K['perda_real_pct'],2)}% real vs meta {br(K.get('perda_meta_pct',0),2)}% "
        f"({'+' if dpf>=0 else ''}{br(dpf,2)} p.p.). Cada 0,1 p.p. de perda em {br(K['moagem_total'])} t equivale a "
        f"{br(K['moagem_total']*0.001)} t de cana.", "vermelho" if dpf>0.2 else "verde")
if "proj_final" in K:
    add(f"PROJECAO fim de safra (ritmo atual {br(K['pace_pct'],0)}%): ~{br(K['proj_final'])} t vs meta cheia "
        f"{br(K['meta_total'])} t = deficit projetado de {br(abs(K['proj_gap']))} t. "
        f"Recuperar exige elevar o ritmo de moagem em {br((K['meta_total']-K['real_periodo'])/max(1,(K['meta_total']-K['meta_periodo']))*100-100 if (K['meta_total']-K['meta_periodo']) else 0,0)}%+ no restante.", "vermelho" if K['proj_gap']<0 else "verde")
if praga.get("broca") is not None:
    add(f"Broca: infestacao media {br(praga['broca'],2)}% (referencia critica ~3%). "
        f"{'Controlado' if praga['broca']<3 else 'ATENCAO - acima do limite'}. Monitorar fazendas-topo.",
        "verde" if praga['broca']<3 else "vermelho")
if plantio.get("aderencia"):
    add(f"Plantio: {br(plantio.get('real_ha',0))} ha realizados vs meta {br(plantio.get('meta_ha',0))} ha = "
        f"{br(plantio['aderencia'],0)}% de aderencia.", "verde" if plantio['aderencia']>=90 else "amarelo")
if "ader_geral" in K and not ader_proc.empty:
    pior = ader_proc.sort_values("ADER").iloc[0]
    add(f"Aderencia de tratos: {br(K['ader_geral'],0)}% geral ({br(K['ader_real_ha'])} de {br(K['ader_prev_ha'])} ha). "
        f"Processo mais atrasado: {pior['PROC']} ({br(pior['ADER'],0)}%). Tratos em dia sustentam o TCH/TAH das proximas safras.",
        "vermelho" if K['ader_geral']<60 else "amarelo")
if manut.get("disp_real") is not None:
    dd = manut["disp_real"]-manut.get("disp_meta",0)
    add(f"DISPONIBILIDADE da frota (medida oficial, mes atual): {br(manut['disp_real'],1)}% vs meta {br(manut.get('disp_meta',0),1)}% "
        f"({'+' if dd>=0 else ''}{br(dd,1)} p.p.). {'ACIMA da meta — frota disponivel' if dd>=0 else 'abaixo da meta'}. "
        f"Logo, o atraso de moagem NAO e por indisponibilidade de frota — o foco e ritmo/clima/fornecedor.",
        "verde" if dd>=0 else "vermelho")
if manut.get("corretiva_pct") is not None:
    add(f"MANUTENCAO {br(manut['corretiva_pct'],0)}% CORRETIVA ({br(manut.get('os',0),0)} OS) — apesar da disponibilidade OK, "
        f"a manutencao reativa e RISCO de custo e de quebra futura. Implementar preventiva protege o ritmo. "
        f"DMT {br(manut.get('dmt',0),1)} km; diesel {br(manut.get('diesel_l',0)/1e6,2)} M L.",
        "amarelo")

# ── ACOES PRIORIZADAS (onde atuar) ───────────────────────────────────────────
acoes = []
def acao(prio, titulo, base, meta, dono): acoes.append({"p":prio,"t":titulo,"b":base,"m":meta,"d":dono})
if K.get("gap_periodo",0) < 0:
    acao("ALTA","Acelerar ritmo de moagem/colheita",
         f"Atraso de {br(abs(K['gap_periodo']))} t vs meta ({br(K.get('atend_periodo',0),0)}% do ritmo); projecao fecha em {br(K.get('proj_final',0))} t",
         "Recuperar o gap ate o fim da safra", "Industrial + CCT (Flavio Faveri)")
_mpatr = K.get("meta_atr_periodo", META_ATR)
if K.get("atr_pond",0) < _mpatr - 1.5:
    acao("ALTA","Elevar ATR via maturacao e ponto de corte",
         f"ATR {br(K['atr_pond'],1)} kg/t vs meta ponderada da data {br(_mpatr,1)} (gap {br(K['atr_pond']-_mpatr,1)})",
         f"Priorizar talhoes maduros + maturador; meta cheia {br(K.get('meta_atr_full',META_ATR),1)} kg/t", "Agricola/Maturacao")
elif K.get("atr_pond",0) < _mpatr:
    acao("MEDIA","Acompanhar ATR (levemente abaixo da meta da data)",
         f"ATR {br(K['atr_pond'],1)} vs meta data {br(_mpatr,1)} (gap {br(K['atr_pond']-_mpatr,1)}) — dentro do esperado p/ a fase",
         "Sustentar curva de maturacao ate o pico set-out", "Agricola/Maturacao")
if not frentes.empty and "Fabiano (F27)" in frentes.set_index("GRP").index:
    acao("MEDIA","Cobrar qualidade do fornecedor Fabiano (F27)",
         f"ATR {br(frentes.set_index('GRP').loc['Fabiano (F27)','ATR'],1)} kg/t, abaixo da propria e da Lerosa",
         "Equiparar ao padrao das frentes proprias", "Suprimentos/Fornecedores")
if manut.get("corretiva_pct",0) > 70:
    acao("MEDIA","Implantar manutencao preventiva da frota",
         f"{br(manut['corretiva_pct'],0)}% das OS sao corretivas (reativa). Disponibilidade atual OK ({br(manut.get('disp_real',0),0)}%), "
         f"mas e risco de custo e de quebra futura",
         "Reduzir corretiva sem perder disponibilidade", "Manutencao Automotiva")
if not paradas.empty:
    acao("MEDIA","Reduzir paradas do maior ofensor",
         f"'{paradas.iloc[0]['GRUPO']}' = {br(paradas.iloc[0]['HORAS'],0)} h",
         "Plano de acao por motivo de parada", "Industrial/Manutencao")
if not cst_grupo.empty:
    est = cst_grupo[cst_grupo['DESVIO']>0].sort_values('DESVIO',ascending=False)
    if not est.empty:
        acao("MEDIA","Conter estouro de custo",
             f"Grupo '{est.iloc[0]['DE_GRUPO']}' +R$ {br(est.iloc[0]['DESVIO'])} vs orcado",
             "Trazer realizado ao orcado", "Controladoria")

print(f"  KPIs={len(K)} insights={len(insights)} acoes={len(acoes)}")
print("[3/4] Gerando HTML...")

# ── HTML ─────────────────────────────────────────────────────────────────────
def kpi_card(label, valor, sub="", cor="verde", delta=""):
    return f"""<div class="kpi">
      <div class="kpi-top"><span class="dot {cor}"></span><span class="kpi-label">{label}</span></div>
      <div class="kpi-val">{valor}</div>
      <div class="kpi-sub">{sub} {('<span class=\"delta '+cor+'\">'+delta+'</span>') if delta else ''}</div>
    </div>"""

cards = []
if "moagem_total" in K:
    ap = K.get("atend_periodo", K.get("atend_meta",0))
    sit = "verde" if ap>=98 else ("amarelo" if ap>=90 else "vermelho")
    cards.append(kpi_card("Moagem realizada", f"{br(K['moagem_total'])} t",
                          f"vs meta pro-rata ate {K.get('data_analise','')}: {br(K.get('meta_periodo',0))} t", sit,
                          f"{br(ap,1)}% do ritmo"))
if "atr_pond" in K and K["atr_pond"]>0:
    mp = K.get("meta_atr_periodo", META_ATR)
    d = K["atr_pond"]-mp
    cards.append(kpi_card("ATR ponderado", f"{br(K['atr_pond'],2)} kg/t",
                          f"meta ate data {br(mp,1)} | safra {br(K.get('meta_atr_full',META_ATR),1)}",
                          "verde" if d>=-1.5 else ("amarelo" if d>=-4 else "vermelho"),
                          f"{'+' if d>=0 else ''}{br(d,2)} kg/t vs meta data"))
if "gap_periodo" in K:
    g = K["gap_periodo"]
    cards.append(kpi_card("Aderencia a meta", f"{'+' if g>=0 else ''}{br(g)} t",
                          f"{'adiantado' if g>=0 else 'atrasado'} vs ritmo", "verde" if g>=0 else "vermelho",
                          f"meta cheia {br(K.get('meta_total',META_MOAGEM)/1e6,2)} M t"))
if "proj_final" in K:
    cards.append(kpi_card("Projecao fim de safra", f"{br(K['proj_final']/1e6,2)} M t",
                          f"no ritmo de {br(K.get('pace_pct',0),0)}%", "vermelho" if K['proj_gap']<0 else "verde",
                          f"deficit {br(K['proj_gap']/1e6,2)} M t"))
if K.get("receita_real"):
    cards.append(kpi_card("Receita ATR (realizada)", f"R$ {br(K['receita_real']/1e6,1)} M",
                          f"CONSECANA {br(PRECO_ATR,2)}/kg", "info",
                          (f"risco proj. {br(K['receita_gap']/1e6,0)} M" if K.get("receita_gap") is not None else "")))
if K.get("perda_real_pct") is not None:
    dpf = K["perda_real_pct"]-K.get("perda_meta_pct",0)
    cards.append(kpi_card("Perda industrial", f"{br(K['perda_real_pct'],2)}%",
                          f"meta {br(K.get('perda_meta_pct',0),2)}%", "vermelho" if dpf>0.2 else "verde",
                          f"{'+' if dpf>=0 else ''}{br(dpf,2)} p.p."))
if manut.get("disp_real") is not None:
    dd = manut["disp_real"]-manut.get("disp_meta",0)
    cards.append(kpi_card("Disponibilidade frota", f"{br(manut['disp_real'],1)}%",
                          f"meta {br(manut.get('disp_meta',0),1)}% (oficial, mes)", "verde" if dd>=0 else "vermelho",
                          f"{'+' if dd>=0 else ''}{br(dd,1)} p.p."))
if manut.get("corretiva_pct") is not None:
    cards.append(kpi_card("Manut. corretiva", f"{br(manut['corretiva_pct'],0)}%",
                          f"{br(manut.get('os',0),0)} OS | DMT {br(manut.get('dmt',0),0)} km", "amarelo",
                          "risco/custo (disp. OK)"))
if manut.get("diesel_l"):
    cards.append(kpi_card("Diesel (frota)", f"{br(manut['diesel_l']/1e6,2)} M L",
                          f"DMT medio {br(manut.get('dmt',0),1)} km", "info"))
if "custo_real" in K:
    dev = (K["custo_real"]-K["custo_orc"])/K["custo_orc"]*100 if K.get("custo_orc") else 0
    cards.append(kpi_card("Custo realizado 2026", f"R$ {br(K['custo_real']/1e6,1)} M",
                          f"orcado R$ {br(K['custo_orc']/1e6,1)} M", "vermelho" if dev>2 else "verde",
                          f"{'+' if dev>=0 else ''}{br(dev,1)}% vs orc"))
if "chuva_ano_atual" in K:
    d = (K["chuva_ano_atual"]/K["chuva_media"]-1)*100 if K.get("chuva_media") else 0
    cards.append(kpi_card("Chuva 2026", f"{br(K['chuva_ano_atual'])} mm",
                          f"media hist. {br(K.get('chuva_media',0))} mm", "info",
                          f"{'+' if d>=0 else ''}{br(d,0)}% vs media"))

def linhas_tabela(df, cols, fmts):
    out = []
    for _, r in df.iterrows():
        tds = "".join(f"<td>{fmts[i](r[c])}</td>" for i,c in enumerate(cols))
        out.append(f"<tr>{tds}</tr>")
    return "".join(out)

ins_html = "".join(
    f'<div class="insight {i["tipo"]}">{i["t"]}</div>' for i in insights
) or '<div class="insight info">Sem dados suficientes para conclusoes.</div>'

_pcor = {"ALTA":"vermelho","MEDIA":"amarelo","BAIXA":"info"}
acoes_html = "".join(
    f'<tr><td><span class="badge {_pcor.get(a["p"],"info")}">{a["p"]}</span></td>'
    f'<td><b>{a["t"]}</b><div class="muted">{a["b"]}</div></td>'
    f'<td>{a["m"]}</td><td>{a["d"]}</td></tr>' for a in sorted(acoes, key=lambda x:{"ALTA":0,"MEDIA":1,"BAIXA":2}.get(x["p"],3))
) or '<tr><td colspan=4>sem acoes</td></tr>'

ader_rows = linhas_tabela(ader_proc, ["PROC","PREV","REAL","ADER"],
                  [str, lambda v:br(v,0), lambda v:br(v,0), lambda v:br(v,0)+"%"]) if not ader_proc.empty else '<tr><td colspan=4>sem dados</td></tr>'
md_rows = linhas_tabela(meta_disp, ["CATEGORIA","META_DISP"],
                  [str, lambda v:br(v*100,1)+"%"]) if not meta_disp.empty else '<tr><td colspan=2>sem dados</td></tr>'
frentes_rows = linhas_tabela(frentes, ["GRP","CANA","META","ADER","ATR"],
                  [str, lambda v:br(v,0), lambda v:br(v,0), lambda v:br(v,0)+"%", lambda v:br(v,1)]) if not frentes.empty else '<tr><td colspan=5>sem dados</td></tr>'

# Tabelas
tah_rows = linhas_tabela(tah_var.head(15), ["DE_VARIED","TAH","TCH","AREA"],
                         [str, lambda v:br(v,2), lambda v:br(v,1), lambda v:br(v,0)]) if not tah_var.empty else '<tr><td colspan=4>sem dados</td></tr>'
cst_rows = linhas_tabela(cst_grupo, ["DE_GRUPO","ORC","REAL","DESVIO_PCT"],
                         [str, lambda v:"R$ "+br(v), lambda v:"R$ "+br(v), lambda v:("+" if v>=0 else "")+br(v,1)+"%"]) if not cst_grupo.empty else '<tr><td colspan=4>sem dados</td></tr>'
atrmat_rows = linhas_tabela(atr_mat, ["DE_MATURAC","ATR","TON"],
                         [str, lambda v:br(v,2), lambda v:br(v,0)]) if not atr_mat.empty else '<tr><td colspan=3>sem dados</td></tr>'
varhist_rows = linhas_tabela(var_hist.head(15), ["VAR","TAH","TCH","AREA"],
                         [str, lambda v:br(v,2), lambda v:br(v,1), lambda v:br(v,0)]) if not var_hist.empty else '<tr><td colspan=4>sem dados</td></tr>'
varamb_rows = linhas_tabela(var_amb.head(15), ["VAR","AMB1","TAH","AREA"],
                         [str, str, lambda v:br(v,2), lambda v:br(v,0)]) if not var_amb.empty else '<tr><td colspan=4>sem dados</td></tr>'

html = f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>UMOE | Cockpit Executivo 2026/27</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
:root{{--bg:#0A0F1E;--surf:#111A30;--surf2:#16213D;--line:#22325C;--txt:#E8EEFC;--mut:#8FA3C8;--green:#22C55E;--gold:#FACC15;--red:#F43F5E;--blue:#3B82F6;--cyan:#22D3EE}}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:var(--bg);color:var(--txt);font-family:'Segoe UI',system-ui,sans-serif;font-size:14px}}
header{{padding:22px 34px;border-bottom:1px solid var(--line);display:flex;align-items:center;justify-content:space-between;background:var(--surf)}}
header h1{{font-size:1.25rem;color:#fff;font-weight:700;letter-spacing:.3px}}
header .sub{{color:var(--mut);font-size:.8rem;margin-top:3px}}
.tag{{background:var(--green);color:#04210f;padding:3px 12px;border-radius:20px;font-size:.72rem;font-weight:700}}
nav{{display:flex;gap:2px;padding:0 28px;background:var(--surf);border-bottom:1px solid var(--line);overflow-x:auto}}
nav button{{background:none;border:0;color:var(--mut);padding:13px 18px;cursor:pointer;font-size:.84rem;font-weight:600;border-bottom:2px solid transparent;white-space:nowrap}}
nav button.on,nav button:hover{{color:#fff;border-bottom-color:var(--gold)}}
.tab{{display:none;padding:26px 34px;max-width:1500px;margin:0 auto}}
.tab.on{{display:block}}
.kpis{{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:16px;margin-bottom:26px}}
.kpi{{background:var(--surf);border:1px solid var(--line);border-radius:14px;padding:18px 20px}}
.kpi-top{{display:flex;align-items:center;gap:8px;margin-bottom:10px}}
.kpi-label{{color:var(--mut);font-size:.74rem;text-transform:uppercase;letter-spacing:.7px;font-weight:600}}
.kpi-val{{font-size:2rem;font-weight:800;color:#fff;line-height:1.1}}
.kpi-sub{{color:var(--mut);font-size:.78rem;margin-top:6px}}
.dot{{width:9px;height:9px;border-radius:50%}}
.dot.verde{{background:var(--green)}} .dot.amarelo{{background:var(--gold)}} .dot.vermelho{{background:var(--red)}} .dot.info{{background:var(--blue)}}
.delta{{font-weight:700}} .delta.verde{{color:var(--green)}} .delta.amarelo{{color:var(--gold)}} .delta.vermelho{{color:var(--red)}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-bottom:18px}}
.grid.one{{grid-template-columns:1fr}}
@media(max-width:900px){{.grid{{grid-template-columns:1fr}}}}
.card{{background:var(--surf);border:1px solid var(--line);border-radius:14px;padding:20px}}
.card h3{{font-size:.82rem;color:var(--gold);text-transform:uppercase;letter-spacing:.8px;margin-bottom:16px;font-weight:700}}
.chart{{position:relative;height:280px}}
table{{width:100%;border-collapse:collapse;font-size:.82rem}}
th{{text-align:left;color:var(--mut);font-size:.7rem;text-transform:uppercase;letter-spacing:.5px;padding:8px 10px;border-bottom:1px solid var(--line)}}
td{{padding:8px 10px;border-bottom:1px solid #16213d;color:var(--txt)}}
tr:hover td{{background:var(--surf2)}}
.insight{{background:var(--surf2);border-left:3px solid var(--blue);border-radius:0 10px 10px 0;padding:13px 16px;margin-bottom:12px;font-size:.9rem;line-height:1.5}}
.insight.verde{{border-left-color:var(--green)}} .insight.amarelo{{border-left-color:var(--gold)}} .insight.vermelho{{border-left-color:var(--red)}}
.badge{{padding:3px 10px;border-radius:20px;font-size:.7rem;font-weight:700}}
.badge.vermelho{{background:rgba(244,63,94,.15);color:#fda4b4}} .badge.amarelo{{background:rgba(250,204,21,.15);color:#fde047}} .badge.info{{background:rgba(59,130,246,.15);color:#93c5fd}}
.muted{{color:var(--mut);font-size:.76rem;margin-top:3px}}
.foot{{color:var(--mut);font-size:.72rem;padding:20px 34px;border-top:1px solid var(--line);text-align:center}}
</style></head><body>
<header>
  <div><h1>UMOE BIOENERGY — Cockpit Executivo</h1><div class="sub">Plano Diretor Agricola | Safra 2026/27 | Dados reais Power BI</div></div>
  <div style="text-align:right"><span class="tag">dados reais ate {K.get('data_analise', HOJE)}</span><div class="sub">Gerado {HOJE} | safra com dados ate {K.get('data_analise','')}</div></div>
</header>
<nav>
  <button class="on" onclick="tab('cockpit',this)">Cockpit</button>
  <button onclick="tab('moagem',this)">Moagem vs Meta</button>
  <button onclick="tab('atr',this)">ATR & Qualidade</button>
  <button onclick="tab('comp',this)">Comparativo Safras</button>
  <button onclick="tab('frentes',this)">Frentes & Fornecedores</button>
  <button onclick="tab('varied',this)">Variedades</button>
  <button onclick="tab('chuva',this)">Chuva & Clima</button>
  <button onclick="tab('pragas',this)">Pragas & Plantio</button>
  <button onclick="tab('tratos',this)">Aderencia Tratos</button>
  <button onclick="tab('manut',this)">Manutencao & Frota</button>
  <button onclick="tab('custos',this)">Custos</button>
  <button onclick="tab('paradas',this)">Disponibilidade</button>
  <button onclick="tab('acoes',this)">Onde Atuar</button>
  <button onclick="tab('intel',this)">Inteligencia</button>
</nav>

<div id="t-cockpit" class="tab on">
  <div class="kpis">{''.join(cards) or '<div class="kpi"><div class="kpi-val">sem dados</div></div>'}</div>
  <div class="grid">
    <div class="card"><h3>Moagem acumulada — realizado vs meta (t)</h3><div class="chart"><canvas id="cAcum"></canvas></div></div>
    <div class="card"><h3>ATR por maturacao (kg/t)</h3><table><tr><th>Maturacao</th><th>ATR</th><th>Cana (t)</th></tr>{atrmat_rows}</table></div>
  </div>
  <div class="card"><h3>Conclusoes (calculadas dos dados)</h3>{ins_html}</div>
</div>

<div id="t-moagem" class="tab">
  <div class="grid one"><div class="card"><h3>Moagem mensal — realizado vs meta (t)</h3><div class="chart"><canvas id="cMetaMes"></canvas></div></div></div>
  <div class="grid one"><div class="card"><h3>Curva acumulada — realizado vs meta (t)</h3><div class="chart"><canvas id="cAcum2"></canvas></div></div></div>
</div>

<div id="t-comp" class="tab">
  <div class="grid one"><div class="card"><h3>Moagem acumulada por safra — no MESMO ponto (ate {K.get('pace_cutoff','')})</h3><div class="chart"><canvas id="cPace"></canvas></div></div></div>
  <div class="grid one"><div class="card"><h3>ATR mensal por safra (kg/t) — curva de maturacao comparada</h3><div class="chart"><canvas id="cAtrMesSaf"></canvas></div></div></div>
  <div class="grid">
    <div class="card"><h3>TAH por safra (t ATR/ha) — base historica 8 safras</h3><div class="chart"><canvas id="cHistTAH"></canvas></div></div>
    <div class="card"><h3>TCH por safra (t/ha) — base historica</h3><div class="chart"><canvas id="cHistTCH"></canvas></div></div>
  </div>
</div>

<div id="t-atr" class="tab">
  <div class="grid one"><div class="card"><h3>ATR — realizado mensal vs meta ponderada (kg/t)</h3><div class="chart"><canvas id="cAtrMeta"></canvas></div></div></div>
  <div class="grid one"><div class="card"><h3>ATR diario realizado (kg/t) — toda a safra</h3><div class="chart"><canvas id="cAtr"></canvas></div></div></div>
  <div class="card"><h3>ATR por maturacao</h3><table><tr><th>Maturacao</th><th>ATR (kg/t)</th><th>Cana (t)</th></tr>{atrmat_rows}</table></div>
</div>

<div id="t-varied" class="tab">
  <div class="grid">
    <div class="card"><h3>TCH real vs estimado por estagio</h3><div class="chart"><canvas id="cTch"></canvas></div></div>
    <div class="card"><h3>Top variedades por TAH — safra atual (t ATR/ha)</h3><table><tr><th>Variedade</th><th>TAH</th><th>TCH</th><th>Area (ha)</th></tr>{tah_rows}</table></div>
  </div>
  <div class="grid">
    <div class="card"><h3>Ranking historico de variedades — 8 safras (t ATR/ha)</h3><table><tr><th>Variedade</th><th>TAH</th><th>TCH</th><th>Area (ha)</th></tr>{varhist_rows}</table></div>
    <div class="card"><h3>Melhor variedade x ambiente — historico</h3><table><tr><th>Variedade</th><th>Amb.</th><th>TAH</th><th>Area (ha)</th></tr>{varamb_rows}</table></div>
  </div>
</div>

<div id="t-chuva" class="tab">
  <div class="grid">
    <div class="card"><h3>Chuva anual (mm) — fonte oficial pluviometrica</h3><div class="chart"><canvas id="cChuvaA"></canvas></div></div>
    <div class="card"><h3>Chuva mensal 2026 (mm)</h3><div class="chart"><canvas id="cChuvaM"></canvas></div></div>
  </div>
  <div class="card"><h3>Paradas climaticas vs disponibilidade</h3><div class="chart"><canvas id="cPar2"></canvas></div></div>
</div>

<div id="t-custos" class="tab">
  <div class="card"><h3>Top 12 grupos de custo — orcado vs realizado</h3>
    <table><tr><th>Grupo</th><th>Orcado</th><th>Realizado</th><th>Desvio</th></tr>{cst_rows}</table></div>
</div>

<div id="t-paradas" class="tab">
  <div class="grid one"><div class="card"><h3>Horas de parada por grupo</h3><div class="chart"><canvas id="cPar"></canvas></div></div></div>
</div>

<div id="t-tratos" class="tab">
  <div class="grid one"><div class="card"><h3>Aderencia de tratos por processo — area realizada vs prevista (ha)</h3><div class="chart"><canvas id="cAder"></canvas></div></div></div>
  <div class="card"><h3>Aderencia por processo</h3><table><tr><th>Processo</th><th>Prevista (ha)</th><th>Realizada (ha)</th><th>Aderencia</th></tr>{ader_rows}</table></div>
</div>

<div id="t-manut" class="tab">
  <div class="grid one"><div class="card"><h3>Disponibilidade da frota — mensal 2025-2026 (medida oficial) vs meta</h3><div class="chart"><canvas id="cDispMes"></canvas></div></div></div>
  <div class="grid one"><div class="card"><h3>Disponibilidade por categoria (mes atual) — piores 12 vs meta</h3><div class="chart"><canvas id="cDispCat"></canvas></div></div></div>
  <div class="grid">
    <div class="card"><h3>Diesel por categoria (litros, 2026)</h3><div class="chart"><canvas id="cDiesel"></canvas></div></div>
    <div class="card"><h3>Top equipamentos por nº de OS de manutencao</h3><div class="chart"><canvas id="cManutTop"></canvas></div></div>
  </div>
  <div class="grid">
    <div class="card"><h3>Metas oficiais de disponibilidade por categoria</h3>
      <table><tr><th>Categoria</th><th>Meta DISP</th></tr>{md_rows}</table></div>
    <div class="card"><h3>Indicadores de frota</h3>
      <table><tr><th>Indicador</th><th>Valor</th></tr>
      <tr><td>OS de manutencao (2026)</td><td>{br(manut.get('os',0),0)}</td></tr>
      <tr><td>% corretiva</td><td>{br(manut.get('corretiva_pct',0),0)}%</td></tr>
      <tr><td>Diesel total</td><td>{br(manut.get('diesel_l',0)/1e6,2)} M L</td></tr>
      <tr><td>DMT (distancia media transporte)</td><td>{br(manut.get('dmt',0),1)} km</td></tr>
      <tr><td>Viagens / Cargas</td><td>{br(manut.get('viagens',0),0)} / {br(manut.get('cargas',0),0)}</td></tr>
      </table></div>
  </div>
</div>

<div id="t-frentes" class="tab">
  <div class="grid">
    <div class="card"><h3>Moagem por frente — Propria x Lerosa x Fabiano (t)</h3><div class="chart"><canvas id="cFrenMoa"></canvas></div></div>
    <div class="card"><h3>ATR por frente (kg/t) — qualidade da materia-prima</h3><div class="chart"><canvas id="cFrenAtr"></canvas></div></div>
  </div>
  <div class="card"><h3>Resumo por frente — realizado vs meta oficial (pro-rata a data)</h3><table><tr><th>Frente</th><th>Realizado (t)</th><th>Meta (t)</th><th>Aderencia</th><th>ATR (kg/t)</th></tr>{frentes_rows}</table></div>
</div>

<div id="t-pragas" class="tab">
  <div class="grid">
    <div class="card"><h3>Broca — top fazendas (% infestacao)</h3><div class="chart"><canvas id="cBroca"></canvas></div></div>
    <div class="card"><h3>Indicadores agronomicos</h3>
      <table><tr><th>Indicador</th><th>Valor</th><th>Referencia</th></tr>
      <tr><td>Broca (infestacao media)</td><td>{br(praga.get('broca',0),2)}%</td><td>critico ~3%</td></tr>
      <tr><td>Cigarrinha (insetos amostrados)</td><td>{br(praga.get('cig_total',0),0)}</td><td>{praga.get('cig_fazendas',0)} fazendas</td></tr>
      <tr><td>Plantio realizado</td><td>{br(plantio.get('real_ha',0),0)} ha</td><td>meta {br(plantio.get('meta_ha',0),0)} ha</td></tr>
      <tr><td>Aderencia de plantio</td><td>{br(plantio.get('aderencia',0),0)}%</td><td>meta 100%</td></tr>
      </table></div>
  </div>
</div>

<div id="t-acoes" class="tab">
  <div class="card"><h3>Onde atuar — plano de acao priorizado (gerado dos indicadores)</h3>
    <table><tr><th>Prioridade</th><th>Acao / diagnostico</th><th>Meta</th><th>Responsavel</th></tr>{acoes_html}</table>
  </div>
  <div class="card"><h3>Conclusoes de suporte</h3>{ins_html}</div>
</div>

<div id="t-intel" class="tab">
  <div class="card"><h3>Inteligencia executiva — conclusoes baseadas em dados reais</h3>{ins_html}</div>
</div>

<div class="foot">UMOE OS 8.0 | Cockpit Executivo | {len(K)} indicadores · dados reais Power BI | calculos alinhados ao DAX oficial (ATR, TCH=PRD/AREA_REEST3, TAH=(ACUCAR/CANA_ACUCAR)x(TON/AREA)/1000) | {HOJE}</div>

<script>
const CT={json.dumps(charts, ensure_ascii=False, default=str)};
const G='#22C55E',O='#FACC15',B='#3B82F6',R='#F43F5E',C='#22D3EE',M='#8FA3C8';
const opt={{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{labels:{{color:M,font:{{size:11}}}}}}}},scales:{{x:{{ticks:{{color:M}},grid:{{color:'#1b2848'}}}},y:{{ticks:{{color:M}},grid:{{color:'#1b2848'}}}}}}}};
function tab(id,b){{document.querySelectorAll('.tab').forEach(t=>t.classList.remove('on'));document.querySelectorAll('nav button').forEach(x=>x.classList.remove('on'));document.getElementById('t-'+id).classList.add('on');b.classList.add('on');}}
// Realizado vs meta acumulada (curva S)
function mkAcum(cid){{const d=CT.meta_mes||[];if(!d.length)return;new Chart(document.getElementById(cid),{{type:'line',data:{{labels:d.map(x=>x.AM),datasets:[
  {{label:'Meta acumulada',data:d.map(x=>x.META_ACUM),borderColor:O,backgroundColor:'transparent',borderDash:[6,4],tension:.2,pointRadius:0}},
  {{label:'Realizado acumulado',data:d.map(x=>x.REAL_ACUM),borderColor:G,backgroundColor:G+'22',fill:true,tension:.2,pointRadius:3}}]}},options:opt}});}}
mkAcum('cAcum'); mkAcum('cAcum2');
// Realizado vs meta mensal
(()=>{{const d=CT.meta_mes||[];if(!d.length)return;new Chart(document.getElementById('cMetaMes'),{{type:'bar',data:{{labels:d.map(x=>x.AM),datasets:[
  {{label:'Realizado',data:d.map(x=>x.REAL),backgroundColor:G+'cc',borderColor:G,borderWidth:1}},
  {{label:'Meta',data:d.map(x=>x.META),backgroundColor:O+'66',borderColor:O,borderWidth:1}}]}},options:opt}});}})();
// Historico por safra
function mkHist(cid,campo,cor){{const d=CT.hist_safra||[];if(!d.length)return;new Chart(document.getElementById(cid),{{type:'bar',data:{{labels:d.map(x=>x.safra),datasets:[{{label:campo,data:d.map(x=>x[campo]),backgroundColor:d.map(x=>x.safra.includes('*')?R:cor+'cc'),borderColor:cor,borderWidth:1}}]}},options:opt}});}}
mkHist('cHistTCH','TCH',G); mkHist('cHistTAH','TAH',C);
// Moagem acumulada por safra no mesmo ponto + ATR (eixo duplo)
(()=>{{const d=CT.moagem_pace||[];if(!d.length)return;new Chart(document.getElementById('cPace'),{{type:'bar',data:{{labels:d.map(x=>x.safra),datasets:[
  {{label:'Moagem acum (t)',data:d.map(x=>x.MOA),backgroundColor:d.map(x=>x.safra.startsWith('26')?G:B+'aa'),borderColor:B,borderWidth:1,yAxisID:'y'}},
  {{label:'ATR (kg/t)',type:'line',data:d.map(x=>x.ATR),borderColor:O,backgroundColor:'transparent',tension:.3,yAxisID:'y1'}}]}},
  options:{{...opt,scales:{{x:{{ticks:{{color:M}},grid:{{color:'#1b2848'}}}},y:{{position:'left',ticks:{{color:M}},grid:{{color:'#1b2848'}}}},y1:{{position:'right',ticks:{{color:O}},grid:{{display:false}}}}}}}}}});}})();
// ATR mensal por safra (multi-linha)
(()=>{{const d=CT.atr_mes_safra;if(!d||!d.meses)return;const cores={{'2022':M,'2023':B,'2024':C,'2025':O,'2026':G}};const nm={{1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}};
  new Chart(document.getElementById('cAtrMesSaf'),{{type:'line',data:{{labels:d.meses.map(m=>nm[m]||m),datasets:Object.keys(d.series).map(s=>({{label:s,data:d.series[s].map(v=>v||null),borderColor:cores[s]||M,backgroundColor:'transparent',borderWidth:s==='2026'?3:1.5,tension:.3,pointRadius:s==='2026'?3:0,spanGaps:true}}))}},options:opt}});}})();
(()=>{{const d=CT.atr_dia||[];if(!d.length)return;new Chart(document.getElementById('cAtr'),{{type:'line',data:{{labels:d.map(x=>x.dia),datasets:[{{label:'ATR kg/t',data:d.map(x=>x.ATR),borderColor:O,backgroundColor:O+'22',fill:true,tension:.35,pointRadius:0}}]}},options:opt}});}})();
// ATR realizado mensal vs meta ponderada
(()=>{{const r=CT.moagem_mes||[],m=CT.meta_atr_mes||[];if(!r.length&&!m.length)return;const labs=(m.length?m:r).map(x=>x.AM||x.MES);
  const rmap={{}};r.forEach(x=>rmap[x.MES||x.AM]=x.ATR);
  new Chart(document.getElementById('cAtrMeta'),{{type:'line',data:{{labels:labs,datasets:[
   {{label:'Realizado',data:labs.map(l=>rmap[l]??null),borderColor:G,backgroundColor:G+'22',fill:true,tension:.3,pointRadius:3,spanGaps:true}},
   {{label:'Meta (ponderada)',data:m.map(x=>x.META_ATR),borderColor:O,backgroundColor:'transparent',borderDash:[6,4],tension:.3,pointRadius:0}}]}},options:opt}});}})();
(()=>{{const d=CT.tch_estagio||[];if(!d.length)return;new Chart(document.getElementById('cTch'),{{type:'bar',data:{{labels:d.map(x=>x.ESTAGIO),datasets:[{{label:'TCH real',data:d.map(x=>x.TCH_REAL),backgroundColor:G+'cc'}},{{label:'TCH estimado',data:d.map(x=>x.TCH_EST),backgroundColor:O+'cc'}}]}},options:opt}});}})();
function mkPar(cid){{const d=CT.paradas||[];if(!d.length)return;new Chart(document.getElementById(cid),{{type:'bar',data:{{labels:d.map(x=>x.GRUPO),datasets:[{{label:'Horas',data:d.map(x=>x.HORAS),backgroundColor:R+'cc',borderColor:R,borderWidth:1}}]}},options:{{...opt,indexAxis:'y'}}}});}}
mkPar('cPar'); mkPar('cPar2');
(()=>{{const d=CT.chuva_ano||[];if(!d.length)return;new Chart(document.getElementById('cChuvaA'),{{type:'bar',data:{{labels:d.map(x=>x.ANO),datasets:[{{label:'Chuva (mm)',data:d.map(x=>x.MM),backgroundColor:C+'aa',borderColor:C,borderWidth:1}}]}},options:opt}});}})();
(()=>{{const d=CT.chuva_mes||[];if(!d.length)return;new Chart(document.getElementById('cChuvaM'),{{type:'bar',data:{{labels:d.map(x=>x.MES),datasets:[{{label:'mm 2026',data:d.map(x=>x.MM),backgroundColor:B+'aa',borderColor:B,borderWidth:1}}]}},options:opt}});}})();
// Frentes: realizado vs meta
(()=>{{const d=CT.frentes||[];if(!d.length)return;const col=d.map(x=>x.GRP.includes('Propria')?G:(x.GRP.includes('Lerosa')?B:O));
  new Chart(document.getElementById('cFrenMoa'),{{type:'bar',data:{{labels:d.map(x=>x.GRP),datasets:[
    {{label:'Realizado (t)',data:d.map(x=>x.CANA),backgroundColor:G+'cc',borderColor:G,borderWidth:1}},
    {{label:'Meta (t)',data:d.map(x=>x.META),backgroundColor:O+'66',borderColor:O,borderWidth:1}}]}},options:opt}});
  new Chart(document.getElementById('cFrenAtr'),{{type:'bar',data:{{labels:d.map(x=>x.GRP),datasets:[{{label:'ATR (kg/t)',data:d.map(x=>x.ATR),backgroundColor:col}}]}},options:opt}});}})();
// Broca top fazendas
(()=>{{const d=CT.broca_top||[];if(!d.length)return;new Chart(document.getElementById('cBroca'),{{type:'bar',data:{{labels:d.map(x=>x.FAZENDA),datasets:[{{label:'% infest',data:d.map(x=>x.INFEST),backgroundColor:R+'cc',borderColor:R,borderWidth:1}}]}},options:{{...opt,indexAxis:'y'}}}});}})();
// Diesel por categoria + manut top
(()=>{{const d=CT.diesel_cat||[];if(!d.length)return;new Chart(document.getElementById('cDiesel'),{{type:'bar',data:{{labels:d.map(x=>x.cat),datasets:[{{label:'Litros',data:d.map(x=>x.L),backgroundColor:O+'cc',borderColor:O,borderWidth:1}}]}},options:{{...opt,indexAxis:'y'}}}});}})();
(()=>{{const d=CT.manut_top||[];if(!d.length)return;new Chart(document.getElementById('cManutTop'),{{type:'bar',data:{{labels:d.map(x=>x.eq),datasets:[{{label:'OS',data:d.map(x=>x.OS),backgroundColor:R+'cc',borderColor:R,borderWidth:1}}]}},options:{{...opt,indexAxis:'y'}}}});}})();
(()=>{{const d=CT.disp_cat||[];if(!d.length)return;new Chart(document.getElementById('cDispCat'),{{type:'bar',data:{{labels:d.map(x=>x.cat),datasets:[
  {{label:'Disponibilidade %',data:d.map(x=>x.disp),backgroundColor:d.map(x=>x.disp>=x.meta?G+'cc':R+'cc')}},
  {{label:'Meta %',type:'line',data:d.map(x=>x.meta),borderColor:O,backgroundColor:'transparent',pointRadius:0}}]}},options:opt}});}})();
(()=>{{const d=CT.ader_proc||[];if(!d.length)return;new Chart(document.getElementById('cAder'),{{type:'bar',data:{{labels:d.map(x=>x.proc),datasets:[
  {{label:'Prevista (ha)',data:d.map(x=>x.prev),backgroundColor:O+'66',borderColor:O,borderWidth:1}},
  {{label:'Realizada (ha)',data:d.map(x=>x.real),backgroundColor:G+'cc',borderColor:G,borderWidth:1}}]}},options:{{...opt,indexAxis:'y'}}}});}})();
(()=>{{const d=CT.disp_mensal||[];if(!d.length)return;new Chart(document.getElementById('cDispMes'),{{type:'line',data:{{labels:d.map(x=>x.ym),datasets:[
  {{label:'Disponibilidade %',data:d.map(x=>x.disp),borderColor:G,backgroundColor:G+'22',fill:true,tension:.3,pointRadius:2}},
  {{label:'Meta %',data:d.map(x=>x.meta),borderColor:O,borderDash:[6,4],backgroundColor:'transparent',pointRadius:0}}]}},options:{{...opt,scales:{{...opt.scales,y:{{...opt.scales.y,min:80,max:100}}}}}}}});}})();
</script></body></html>"""

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(html, encoding="utf-8")
DOCS.mkdir(exist_ok=True)
(DOCS / "UMOE_Cockpit_Executivo.html").write_text(html, encoding="utf-8")
print(f"[4/4] OK -> {OUT} ({len(html)//1024} KB) + docs/")
