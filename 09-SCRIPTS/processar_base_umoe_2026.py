"""
processar_base_umoe_2026.py
Plano Diretor Agricola UMOE -- Validacao e integracao do arquivo
BASE_DE_DADOS_UMOE_2026.xlsx ao pipeline TOPO_

Origem do dado: Geotecnologia / SIG -- mapeamento de solo e topografia
233 fazendas mapeadas em PDF -> consolidado nesta planilha mestre

Uso: python processar_base_umoe_2026.py
"""

import re
import logging
from datetime import datetime
from pathlib import Path
from collections import Counter, defaultdict

import pandas as pd
import openpyxl

# -- CONFIGURACAO --------------------------------------------------------------

INPUT_DIR  = Path(r"C:\Users\andrei.elastico\Downloads\PlanoFazendas_Inputs")
OUTPUT_DIR = INPUT_DIR / "outputs"

ARQUIVO_ORIGEM = INPUT_DIR / "BASE_DE_DADOS_UMOE_2026.xlsx"
# Se o arquivo baixado mantiver sufixo do navegador, tentar variantes
VARIANTES_NOME = [
    "BASE_DE_DADOS_UMOE_2026.xlsx",
    "BASE DE DADOS UMOE_2026.xlsx",
    "BASE_DE_DADOS_UMOE_2026 (1).xlsx",
    "BASE_DE_DADOS_UMOE_2026__1_.xlsx",
]

DATA_HORA = datetime.now().strftime("%Y%m%d_%H%M")
LOG_FILE  = OUTPUT_DIR / f"log_validacao_TOPO_base_umoe_{DATA_HORA}.txt"

log = logging.getLogger(__name__)


def configurar_logging():
    OUTPUT_DIR.mkdir(exist_ok=True, parents=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(LOG_FILE, encoding="utf-8"),
            logging.StreamHandler()
        ]
    )

# -- MAPEAMENTO DE COLUNAS -- ORIGEM -> PADRAO CLAUDE.md ----------------------

MAPA_COLUNAS = {
    "CodFaz":      "CD_FAZENDA",
    "Talhao":      "CD_TALHAO",
    "Nome_Fazenda":"DE_NOME_FAZENDA",
    "Municipio":   "DE_MUNICIPIO",
    "Proprietario":"DE_PROPRIETARIO",
    "Bloco":       "CD_BLOCO",
    "AreaAGPL":    "AREA_AGPL_HA",
    "AreaOutros":  "AREA_OUTROS_HA",
    "AreaProdut":  "AREA_PRODUTIVA_HA",
    "AreaTalhao":  "AREA_TALHAO_HA",
    "Coord_X":     "COORD_X",
    "Coord_Y":     "COORD_Y",
    "AMBIENTE":    "DE_AMBIENTE_PROD",
    "SOLO":        "CD_SOLO",
    "CLASSE":      "CLASSE_SOLO",
    "QUIMICA":     "QUIMICA_SOLO",
    "TEXTURA":     "TEXTURA_SOLO",
    "CARACTERIS":  "CARACTERISTICA_SOLO",
    "DESC_SOLO":   "DESC_SOLO_COMPLETA",
    "OBSERVACAO":  "OBSERVACAO",
    "TERRACO":     "FG_TERRACO",
    "%TERRACO":    "PCT_TERRACO",
    "Decliv":      "FAIXA_DECLIVIDADE",
    "MEC":         "STATUS_MECANIZACAO",
    "Layer":       "CD_LAYER_ORIGEM",
}

COLUNAS_OBRIGATORIAS_SAIDA = ["CD_FAZENDA", "CD_TALHAO"]

# Conversor de faixa de declividade em ponto medio numerico (para calculo ponderado)
FAIXA_DECLIV_NUMERICO = {
    "0 - 5%":   2.5,
    "5 - 12%":  8.5,
    "12 - 20%": 16.0,
    "12-20%":   16.0,
    "12–20%": 16.0,  # en-dash U+2013
    ">20%":     25.0,
}

# -- FUNCOES DE VALIDACAO -----------------------------------------------------

def localizar_arquivo() -> Path:
    for nome in VARIANTES_NOME:
        caminho = INPUT_DIR / nome
        if caminho.exists():
            log.info(f"Arquivo localizado: {caminho.name}")
            return caminho
    candidatos = list(INPUT_DIR.glob("*UMOE*2026*.xlsx")) + list(INPUT_DIR.glob("*UMOE_2026*.xlsx"))
    if candidatos:
        log.warning(f"Nome exato nao encontrado. Usando candidato: {candidatos[0].name}")
        return candidatos[0]
    raise FileNotFoundError(
        f"Nenhum arquivo BASE_DE_DADOS_UMOE_2026 encontrado em {INPUT_DIR}"
    )


def ler_base_original(caminho: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    aba = "BASE UMOE" if "BASE UMOE" in wb.sheetnames else wb.sheetnames[0]
    log.info(f"Lendo aba: {aba}")

    ws = wb[aba]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    data = rows[1:]

    df = pd.DataFrame(data, columns=header)
    log.info(f"Linhas lidas: {len(df)} | Colunas: {len(df.columns)}")
    return df


def validar_estrutura(df: pd.DataFrame) -> dict:
    resultado = {}

    resultado["total_linhas"] = len(df)
    resultado["total_colunas"] = len(df.columns)
    resultado["fazendas_unicas"] = df["CodFaz"].nunique()
    resultado["talhoes_unicos"] = len(df)

    resultado["nulos_codfaz"] = df["CodFaz"].isna().sum()
    resultado["nulos_talhao"] = df["Talhao"].isna().sum()
    resultado["nulos_area"] = df["AreaProdut"].isna().sum()
    resultado["nulos_ambiente"] = df["AMBIENTE"].isna().sum()

    chave = df["CodFaz"].astype(str) + "_" + df["Talhao"].astype(str)
    resultado["duplicatas_chave"] = chave.duplicated().sum()

    areas_validas = pd.to_numeric(df["AreaProdut"], errors="coerce")
    resultado["area_total_ha"] = areas_validas.sum()

    resultado["dist_ambiente"] = df["AMBIENTE"].value_counts(dropna=False).to_dict()
    resultado["dist_municipio"] = df["Municipio"].value_counts().to_dict()

    sem_amb = df[df["AMBIENTE"].isna()]
    resultado["fazendas_sem_ambiente"] = (
        sem_amb.groupby("CodFaz")
        .agg(Nome=("Nome_Fazenda", "first"), Talhoes_Sem_Amb=("Talhao", "count"))
        .reset_index()
        .sort_values("Talhoes_Sem_Amb", ascending=False)
    )

    pct_terr = pd.to_numeric(
        df["%TERRACO"].astype(str).str.replace("%", "").str.replace(",", "."),
        errors="coerce"
    )
    resultado["terraco_invalido"] = df[(pct_terr > 100) | (pct_terr < 0)]

    return resultado


def normalizar_colunas(df: pd.DataFrame) -> pd.DataFrame:
    df = df.rename(columns=MAPA_COLUNAS)

    df["DECLIVIDADE_PCT_MEDIO"] = df["FAIXA_DECLIVIDADE"].map(FAIXA_DECLIV_NUMERICO)
    df["FG_AMBIENTE_AUSENTE"] = df["DE_AMBIENTE_PROD"].isna()

    pct = pd.to_numeric(
        df["PCT_TERRACO"].astype(str).str.replace("%", "").str.replace(",", "."),
        errors="coerce"
    )
    df["PCT_TERRACO_VALIDADO"] = pct.where((pct >= 0) & (pct <= 100), other=pd.NA)

    return df


def exportar_relatorio(df_original: pd.DataFrame, df_normalizado: pd.DataFrame,
                       validacao: dict):
    saida_topo = INPUT_DIR / "TOPO_Base_Dados_UMOE_2026_v1.xlsx"

    with pd.ExcelWriter(saida_topo, engine="openpyxl") as writer:
        df_normalizado.to_excel(writer, sheet_name="DADOS", index=False)

        resumo = pd.DataFrame([
            {"Metrica": "Total de talhoes",                      "Valor": validacao["total_linhas"]},
            {"Metrica": "Fazendas unicas",                       "Valor": validacao["fazendas_unicas"]},
            {"Metrica": "Area total produtiva (ha)",             "Valor": round(validacao["area_total_ha"], 1)},
            {"Metrica": "Talhoes sem AMBIENTE",                  "Valor": validacao["nulos_ambiente"]},
            {"Metrica": "Fazendas afetadas (sem ambiente)",      "Valor": len(validacao["fazendas_sem_ambiente"])},
            {"Metrica": "Duplicatas de chave (Fazenda+Talhao)",  "Valor": validacao["duplicatas_chave"]},
            {"Metrica": "Registros %Terraco invalidos",          "Valor": len(validacao["terraco_invalido"])},
        ])
        resumo.to_excel(writer, sheet_name="RESUMO", index=False)

        dist_amb = pd.DataFrame(
            list(validacao["dist_ambiente"].items()),
            columns=["Ambiente", "Talhoes"]
        ).sort_values("Ambiente", na_position="last")
        dist_amb.to_excel(writer, sheet_name="DIST_AMBIENTE", index=False)

        dist_mun = pd.DataFrame(
            list(validacao["dist_municipio"].items()),
            columns=["Municipio", "Talhoes"]
        ).sort_values("Talhoes", ascending=False)
        dist_mun.to_excel(writer, sheet_name="DIST_MUNICIPIO", index=False)

        if not validacao["fazendas_sem_ambiente"].empty:
            validacao["fazendas_sem_ambiente"].to_excel(
                writer, sheet_name="ACAO_SEM_AMBIENTE", index=False
            )

        if not validacao["terraco_invalido"].empty:
            cols_terr = ["CodFaz", "Nome_Fazenda", "Talhao", "%TERRACO"]
            validacao["terraco_invalido"][cols_terr].to_excel(
                writer, sheet_name="ACAO_TERRACO_INVALIDO", index=False
            )

    log.info(f"Arquivo padronizado gerado: {saida_topo}")

    relatorio_txt = OUTPUT_DIR / f"VAL_TOPO_Base_UMOE_2026_{DATA_HORA}.txt"
    with open(relatorio_txt, "w", encoding="utf-8") as f:
        f.write("=" * 70 + "\n")
        f.write("RELATORIO DE VALIDACAO -- BASE_DE_DADOS_UMOE_2026\n")
        f.write(f"Gerado em: {DATA_HORA}\n")
        f.write("=" * 70 + "\n\n")

        status = "APROVADO COM RESSALVAS" if validacao["nulos_ambiente"] > 0 else "APROVADO"
        f.write(f"STATUS: {status}\n\n")

        f.write("--- ESTRUTURA ---\n")
        f.write(f"Total de talhoes:        {validacao['total_linhas']}\n")
        f.write(f"Fazendas unicas:         {validacao['fazendas_unicas']}\n")
        f.write(f"Area total produtiva:    {validacao['area_total_ha']:.1f} ha\n\n")

        f.write("--- QUALIDADE DOS DADOS ---\n")
        f.write(f"Talhoes sem CD_FAZENDA:  {validacao['nulos_codfaz']}\n")
        f.write(f"Talhoes sem CD_TALHAO:   {validacao['nulos_talhao']}\n")
        f.write(f"Talhoes sem Area:        {validacao['nulos_area']}\n")
        f.write(f"Talhoes sem AMBIENTE:    {validacao['nulos_ambiente']} ")
        f.write(f"({len(validacao['fazendas_sem_ambiente'])} fazendas afetadas)\n")
        f.write(f"Duplicatas de chave:     {validacao['duplicatas_chave']}\n")
        f.write(f"%Terraco invalido (>100 ou <0): {len(validacao['terraco_invalido'])} registros\n\n")

        f.write("--- DISTRIBUICAO POR AMBIENTE ---\n")
        for amb, n in sorted(validacao["dist_ambiente"].items(), key=lambda x: str(x[0])):
            amb_label = amb if amb else "SEM AMBIENTE"
            pct = n / validacao["total_linhas"] * 100
            f.write(f"  {amb_label}: {n} talhoes ({pct:.1f}%)\n")

        f.write("\n--- FAZENDAS PRIORITARIAS SEM AMBIENTE (top 10) ---\n")
        for _, row in validacao["fazendas_sem_ambiente"].head(10).iterrows():
            f.write(f"  {row['CodFaz']} -- {row['Nome']}: {row['Talhoes_Sem_Amb']} talhoes\n")

        f.write("\n--- CAMPOS AUSENTES vs. CLAUDE.md ---\n")
        f.write("  SAFRA       -> nao consta. Sera adicionado via cruzamento com COA_PIMS\n")
        f.write("  FORNECEDOR  -> nao consta. Sera adicionado via cruzamento com COA_PIMS\n")
        f.write("  KM_CTT      -> nao consta. Sera adicionado via cruzamento com COA_PIMS\n")

        f.write("\n--- ACOES RECOMENDADAS ---\n")
        f.write(f"1. Solicitar AMBIENTE para {len(validacao['fazendas_sem_ambiente'])} fazendas (ver aba ACAO_SEM_AMBIENTE)\n")
        f.write(f"2. Corrigir {len(validacao['terraco_invalido'])} registros de %Terraco invalido\n")
        f.write("3. Cruzar com COA_PIMS_SF2627_v1.xlsx (Everton) para SAFRA/FORNECEDOR/KM_CTT/TCH/ATR\n")
        f.write(f"4. Arquivo padronizado pronto em: {saida_topo.name}\n")

    log.info(f"Relatorio de validacao gerado: {relatorio_txt}")

    print(f"\n{'='*60}")
    print(f"  VALIDACAO CONCLUIDA -- STATUS: {status}")
    print(f"{'='*60}")
    print(f"  Arquivo padronizado: {saida_topo}")
    print(f"  Relatorio:           {relatorio_txt}")
    print(f"  Log completo:        {LOG_FILE}")
    print(f"{'='*60}\n")


# -- ENTRY POINT --------------------------------------------------------------

def main():
    configurar_logging()
    log.info("=" * 60)
    log.info("VALIDACAO E INTEGRACAO -- BASE_DE_DADOS_UMOE_2026")
    log.info("=" * 60)

    try:
        caminho = localizar_arquivo()
    except FileNotFoundError as e:
        log.error(str(e))
        print(f"\nERRO: {e}")
        print(f"   Verifique se o arquivo esta em: {INPUT_DIR}")
        return

    df_original = ler_base_original(caminho)
    validacao = validar_estrutura(df_original)
    df_normalizado = normalizar_colunas(df_original)
    exportar_relatorio(df_original, df_normalizado, validacao)

    log.info("Processamento concluido.")


if __name__ == "__main__":
    main()
